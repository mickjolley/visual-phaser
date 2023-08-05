# -*- coding: utf-8 -*-

"""
Visual_Phaser.py has the option of adding chromosome 23 (X) to the 
selected chromosomes. 

@author: mick jolley (email: mickj1948@gmail.com)

Visual_Phaser automatically downloads a One-to-One Autosomal match between
individuals. The results are pasted into an .xlsx file which can be read by
Excel 2010 or Libre Office. The result is in a format which allows to to start
manual phasing immediately. No additional manipulations are required.

Any number of individals may be selected.

All 22 chromosomes may be downloaed at once. Alternatively, chromosomes to be
downloaded may be individually selected.

Chromosome 23 can be added to the list of selected chromosomes. In addition,
chromosome 23 can be selected alone.

The length of the downloaded match images can be varied as desired. The default
value of '1' results in the natural size.

Download and install Chrome or Firefox before running this.

reCAPTCHA management.
Press the <Enter> key when encountering a reCAPTCHA message. A short tone
alerts the user. If the window does not immediately change to the One-to-One
comparison page and turns opaque, scroll down to the 'VERIFY' captcha and start
clickingon traffic lights or whatever. Resist the urge to minimize the window 
to take a peek at what the program is doing (the answer is "nothing"). If you
do this make sure that you maximize it before the reCAPTCHA time-out expires,
otherwise the program will crash because it's trying to minimize an already 
minimized window. 

reCAPTCHA time-out.
The default time-out is 30 seconds. This can be changed to whatever you desire.

The operating variables are stored in the VPconfig_V3_0.py file.

You must install the pillow, openpyxl and selenium modules before running this
program.

It is suggested that you creat a new folder to store Visual_Phaser.py
and the .xlsx files generated.

Change in Visual_Phaser.py : AUTO_REC_PNTS is disabled if all the kits 
are phased kits. 

©2023 Michael E. Jolley

"""
import sys
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border,Side,Alignment,PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import urllib.request
from PIL import Image 
import time
import winsound
from itertools import combinations
# import matplotlib.pyplot as plt
import matplotlib.image as mpimg
from openpyxl.utils import get_column_letter
import numpy as np


from VPconfig import *

# Initialize variables.
def initialize():
    
    global match_pair_ids, match_pair_names, steps, ws, wb, driver, url
    global wdir, imtime, xscale, xlname, selected_chromosomes, names, unphflag
    
    # kits = []
    names = [] 
    ids = []
    
    # Input chromosome to analyze
    if len(CHROMOSOMES_SELECTED) == 0:
        selected_chromosomes = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22]
    else:
        selected_chromosomes = CHROMOSOMES_SELECTED
                
    # Extract names and IDs
    for i, item in enumerate(INDIVS):
        
        x = item.split('_')
        names.append(x[0])
        ids.append(x[1])
    
    # Check that kits are all phased
    unphflag = False    
    for name in names:
        if name[-1] != 'P':
            unphflag = True
            break
        
    match_pair_names = list(combinations(names, 2))
    match_pair_ids = list(combinations(ids, 2))
    
    if PHASED_KITS:
    
        match_pair_names = [item for item in match_pair_names if (item[0][-1] == 'P' and item[1][-1] == 'P') or \
           (item[0][-1] != 'P' and item[1][-1] != 'P')]
            
        ph =  ids[-1][-2]    
    
        match_pair_ids = [item for item in match_pair_ids if (item[0][-2] == ph and item[1][-2] == ph) or \
           (item[0][-2] != ph and item[1][-2] != ph)]
        
    # Initialize variables
    xlfilnme = EXCEL_FILE_NAME
    wdir = WORKING_DIRECTORY
    if wdir[-1] != '/':
        wdir = wdir + '/'
    wdir.replace("\\", "/")
    imtime = IMAGE_GATHERING_TIMEOUT
    xscale = CROMOSOME_LENGTH_SCALE_FACTOR 
    
    # Get Excel file name.
    xlpath = wdir + xlfilnme
    xlname = xlpath + '.xlsx'
    
    # Get browser.
    if BROWSER == 'Chrome':
        driver = webdriver.Chrome()
    elif BROWSER == 'Firefox':
        driver = webdriver.Firefox()    
    else:
        print("'Chrome' or 'Firefox' need to be entered.\n Please check browser entry.")
        sys.exit()
    # URL of website
    url = "https://www.gedmatch.com/login1.php"

    # Check if Excel file exists. If so, delete it.
    if os.path.exists(xlname):
        os.remove(xlname)
    
    # Create an new Excel file and delete 'Sheet'.
    try:
        wb = Workbook()
        ws = wb.active
        del wb['Sheet']
        # ws = wb.create_sheet('tables')

    except:
        print("There's a problem with creating this workbook.\nCheck the name")
        sys.exit()


    steps = 1
 
def login(url): 
   
    # Opening the website
    driver.get(url)
    
    # Get email and password locations.
    email = driver.find_element(By.ID, 'email')
    password = driver.find_element(By.NAME, 'password')  
    
    # Send email and password.
    try:    
        email.send_keys(EMAIL)
    except:
        print("There's a problem with the email address.\nPlease check that it was entered correctly.")
        driver.close()
        sys.exit()
    try:
        password.send_keys(PASSWORD)  
    except:
        print("There's a problem with the password.\nPlease check that it was entered correctly.")
        driver.close()
        sys.exit()
        
    # Log in.    
    driver.find_element(By.XPATH,"//*[@id='form-register']/div[2]/button").click()

    # Select "One-to-One Autosomal DNA Comparison".
    driver.find_element(By.PARTIAL_LINK_TEXT,'One-to-One Autosomal').click()


def multisheets(src, src2):
    global er
    # Create worksheets
    for i in range(1, 23):
        ws = wb.create_sheet('Chr{}'.format(i))
        format_sheet(ws)

    # Initialize missing images counter
    tbl = 0
    
    # Get a list of tables
    ch = []
    
    text = ' '
    
    # Check to see if selected chromosomes are in the data
    for n, chrom in enumerate(selected_chromosomes):
        n += 1
        text = ' '
        try:        
            text = driver.find_element(By.XPATH, '//*[@id="Chromosomes"]/table[{}]/tbody/tr[2]/td[1]'.format(chrom)).text
        except:
            pass
        
        # Ignore blank data
        if text == "" :
            print('TEXT ERROR IN CHROMOSOME str(chrom)\n')
            er += 1
            continue
            
        ch.append(text)
        
    # Remove duplicates. Missing FIRs will generate duplicates
    chroms = [*set(ch)]
    
    
    # Indiviuals' name combination
    iname = match_pair_names[q][0] + '-' + match_pair_names[q][1]

    # Cell formatting
    side = Side(border_style='thin')
    border = Border(left=side, right=side, top=side, bottom=side)
    align = Alignment(horizontal='center')

    # Create temporary image files for chromosomes and bars
    for i, chrom in enumerate(selected_chromosomes):
        
        # Create temporary image files during HIR pass
        if steps == 1:
            
            ws = wb['Chr{}'.format(chrom)]
            
            # Set names column width
            if len(iname) >  ws.column_dimensions['J'].width:
                ws.column_dimensions['J'].width = len(iname) + 2
    
            try:
                # Get chromosome images
                urllib.request.urlretrieve(src[20 * (chrom - 1)],"gfg")      
            except Exception as err:
                
                print(err, '\n')
                
                # Paste images to sheets
                fill_sheets(q)
                
                # Add Bold Borders
                if not AUTO_REC_PNTS:
                    add_borders()
                   
                # Auto RP
                if AUTO_REC_PNTS:
                    print('Calculating Recombination Points\n')
                    auto_rp(q)
                
                # Delete empty sheets
                delete_sheets()
                
                # Save workbook.
                wb.save(xlname) 

                # Delete temporary images
                delete_images()
                
                # Close browser.        
                driver.close()

                print('\nConnection Failure. Data saved in ' + xlname + '\n')

                sys.exit()
               
            with Image.open('gfg') as im:
                  for j in range(1):
                    im.seek(im.n_frames // 1 * j)
                    try:
                        im.save(wdir + 'temp{0}{1}.png'.format(q, chrom))
                    except:
                        print("There's a problem with the working directory.\nPlease check that it was entered correctly.")
                        driver.close()
                        sys.exit()
                        
            print('Chromosome data for chromosme {0} acquired for {1} match\n'.format(chrom, iname))
               
    
            # Get bar images
            urllib.request.urlretrieve(src2[15 * (chrom - 1)],"gfg")
            
            with Image.open('gfg') as im:
                  for j in range(1):
                    im.seek(im.n_frames // 1 * j)
                    try:
                        im.save(wdir + 'tempa{0}{1}.png'.format(q, chrom))
                    except:
                        print("There's a problem with the working directory.\nPlease check that it was entered correctly.")
                        driver.close()
                        sys.exit()  
                    
            print('Bar data for chromosme {0} acquired for {1} match\n'.format(chrom, iname))
           
       
    # # Paste tables
    for i, chrom in enumerate(selected_chromosomes):

        if (str(chrom) in chroms):
            
            
            rows = driver.find_elements(By.XPATH, '//*[@id="Chromosomes"]/table[{}]/tbody/tr'.format(chrom - tbl))
            
            nrows = len(rows)
            # Initialize text variable
            text = ''
            
            # Active appropriate sheet
            ws = wb['Chr{}'.format(chrom)]
            
            # FIR table
            if steps == 2:
                next_line = len(ws['B']) + 2
                ws.cell(next_line, 2).value = iname + ' ' + 'FIR Table'
                
            
            elif q == 0:
                format_sheet(ws)
                next_line = 3
                ws.cell(next_line, 2).value = iname
    
            else:   
                next_line = len(ws['B']) + 2
                ws.cell(next_line, 2).value = iname
          
    
            for r in range(1, nrows + 1):
                for p in range(1, 9):
                    text = driver.find_element(By.XPATH,
                        '//*[@id="Chromosomes"]/table[{0}]/tbody/tr[{1}]/td[{2}]'.format(chrom - tbl,r,p)).text    
                    ws.cell(next_line + r , p + 1 ).border = border
                    ws.cell(next_line+ r , p + 1 ).alignment = align
                    ws.cell(r + next_line, p + 1 ).value = text
                        
            print('Table {}, match {} pasted\n'.format(chrom, iname))
            
    
        else:
            print('Table {} has no FIR\n'.format(chrom))
            # Incremnt missing images counter
            tbl += 1
        

def scraper():
    global src, recapno, src2
    # Prevent Hard Breaks.
    driver.find_element(By.XPATH,'//*[@id="NextPage"]/div[11]/input').click()
  
    # Click FIR Check Box
    if steps == 2:
        driver.find_element(By.XPATH,'//*[@id="NextPage"]/div[10]/input').click()
    
    # Set cM cutoof
    driver.find_element(By.XPATH, '//*[@id="NextPage"]/div[8]/input').clear()
    driver.find_element(By.XPATH, '//*[@id="NextPage"]/div[8]/input').send_keys(CM_CUTOFF)

    
    # Send Individuals' GEDmatch ID.
    try:
        driver.find_element(By.XPATH,'//*[@id="NextPage"]/div[1]/input').send_keys(match_pair_ids[q][0])
    except:
        print("There's a problem with Individual1's GEDmatch ID.\n Please check that it was entered correctly.")
        driver.close()
        sys.exit()
    try:
        driver.find_element(By.XPATH,'//*[@id="NextPage"]/div[2]/input').send_keys(match_pair_ids[q][1])
    except:
        print("There's a problem with Individual2's GEDmatch ID.\n Please check that it was entered correctly.")
        driver.close()
        sys.exit() 
            
    # Click "COMPARE". If reCAPTCHA present try manual clicking.
    try:
        driver.find_element(By.XPATH, '//*[@id="NextPage"]/input[2]').click()
    except:
        winsound.Beep(BEEP_TONE, BEEP_TIME)
        driver.maximize_window()
        recapno += 1    
        
        time.sleep(RECAPTCHA_TIMEOUT)

    driver.minimize_window()
    
    results = []
    results2 = []
    
    # Download data
    try:
        results = WebDriverWait(driver, imtime).until(EC.presence_of_all_elements_located((By.XPATH,'//*[@id="Chromosomes"]/div/img')))
        results2 = WebDriverWait(driver, imtime).until(EC.presence_of_all_elements_located((By.XPATH,'//*[@id="Chromosomes"]/img')))
    except :
        # er += 1
        print('Internet failure\n')
        # continue
    finally:
        print( "The data download was successful unless otherwise noted\n")
        
    # Place all source addresses into lists "src" and "src2".
    src = []
    src2 = []
    for img in results:
        src.append(img.get_attribute('src'))
    for img in results2:
        src2.append(img.get_attribute('src'))
    
    return src, src2

def elapsed_time():
    et = time.time()    
    tt = (et - st)/60
    tt = round(tt, 2)
    print('Elapsed time = {} minutes\n'.format(tt))

def delete_images():
    files = os.listdir(wdir)
    for images in files:
        if images.endswith(".png"):
            os.remove(wdir + images)
    
def fill_sheets(matches): 
    for w in range(matches):

        ofs = w * 3

        iname = match_pair_names[w][0] + '-' + match_pair_names[w][1]

        for _,chrom in enumerate(selected_chromosomes):
            # i += 1
            ws = wb['Chr{}'.format(chrom)]
            img = openpyxl.drawing.image.Image(wdir + 'temp{0}{1}.png'.format(w, chrom))
            correction_factor = 1000/img.width
            img.width = correction_factor * img.width * xscale
            img.height = img.height * 20
            ws.add_image(img, ws.cell(ofs + 3, 11).coordinate)
            ws.cell(ofs + 3, 10).value = iname
            ws.cell(ofs + 3, 10).alignment = Alignment(horizontal='center')

        for _,chrom in enumerate(selected_chromosomes):
            # i += 1
            ws = wb['Chr{}'.format(chrom)]
            img = openpyxl.drawing.image.Image(wdir + 'temp{0}{1}.png'.format(w, chrom))
            correction_factor = 1000/img.width
            img = openpyxl.drawing.image.Image(wdir + 'tempa{0}{1}.png'.format(w, chrom))
            img.width = correction_factor * img.width * xscale
            img.height = img.height * 15 
            ws.add_image(img, ws.cell(ofs + 4, 11).coordinate)

def format_sheet(ws):
    ws.column_dimensions['A'].width = 1
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 17 
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 17
    ws.column_dimensions['H'].width = 11
    ws.column_dimensions['I'].width = 16

def delete_sheets():
    for _, name in enumerate(wb.sheetnames):
        ws = wb[name]
        if ws.cell(3, 10).value:
            continue
        else:
            del wb[name]

def add_borders():
    for _, chrom in enumerate(selected_chromosomes):
        ws = wb['Chr{}'.format(chrom)]
        side = Side(border_style='thick')
        border = Border(left=side)
        for i in range(1, 100):
            for j in range(11, 40):
                ws.cell(i, j).border = border
                    
def col_borders_thick(ws, i, j):
    side = Side(border_style='thick')
    border = Border(left=side)
    for n in range(1, j):
        ws.cell(n,i).border = border
                    
def auto_rp(mps, names):
    
    for _, chrom in enumerate(selected_chromosomes):
        
        ws = wb['Chr{}'.format(chrom)]
        
        # Initialize list of recombination points.
        recombs = []    

        for w in range(mps):
            
            # Read Image
            img = mpimg.imread(wdir + 'temp{0}{1}.png'.format(w, chrom))
            
            if match_pair_names[w][0][-1] != 'P' and match_pair_names[w][1][-1] != 'P':
    
                # Separate red and green channels.        
                red = img[0,:,0]
                green = img[0,:,1]
                
                pix1 = red.size
                pix2 = green.size
        
                # Put red channel into list and fill single pixel gaps. Note that red
                # channel = 0 corresponds to green in image.
                rpix = []
                for item in red:
                    rpix.append(item)
                for i, item in enumerate(rpix):
                    if i >= pix1 - 2:
                        break
                    elif rpix[i] == 0 and rpix[i+1] !=0 and rpix[i+2] == 0:
                        rpix[i+1] = 0
                
                # Set x axis for plots.
                x = np.linspace(1,pix1,num = pix1)
                
                gpix = []
                for item in green:
                    gpix.append(item)
                
                # Plot red channel data.
                # plt.scatter(x,rpix,marker='|')
                # plt.show()
                
                # Initialize lists.
                rpr = [0]
                recpr = [0]
                rpg = [0]
                recpg = [0]
        
                # Place green pixels into list rpg.
                for i in range(pix1):
                    if rpix[i] == 0:
                        rpg.append(i)
                        
                rpg.append(pix1)  
                      
                # print(rpg) 
                
                # Algorithm for determining green recombination points. Place these in list
                # recpg.
                for i in range(1, len(rpg)-1):
                    if (rpg[i] < 5 or pix1 - rpg[i] < 5):
                        continue
                    if (rpg[i+1] - rpg[i] >1 and rpg[i] - rpg[i-1] == 1 and\
                        rpg[i] - rpg[i-2] == 2 and rpg[i] - rpg[i-3] == 3 and\
                        rpg[i] - rpg[i-4] == 4 and rpg[i] - rpg[i-5] == 5) or\
                        (rpg[i] - rpg[i-1] >1 and rpg[i+1] - rpg[i] == 1 and\
                         rpg[i+2] - rpg[i] == 2 and rpg[i+3] - rpg[i] == 3 and\
                         rpg[i+4] - rpg[i] == 4 and rpg[i+5] - rpg[i] == 5):    
                            recpg.append(rpg[i])  
            
                            
                recpg.append(pix1)  
                        
                # print('recpg')   
                # print(recpg)  
                # print()
                
                # Plot green channel. Values of <0.3 correspond to red in image.
                # plt.scatter(x,gpix,marker='|')
                # plt.show()
                
            
                # Place red pixels into list rpr.
                for i in range(pix1):
                    if green[i] < 0.3:
                        rpr.append(i)
                        
                rpr.append(pix1) 
                        
                # print(rpr) 
        
                #  Algorithm for determining red recombination points.
                for i in range(1, len(rpr)-1):
                    if (rpr[i] - rpr[i-1] >= NIR_GAP) and (rpr[i+1] - rpr[i] <= NIR_GAP)\
                        or (rpr[i] - rpr[i-1] <= NIR_GAP) and (rpr[i+1] - rpr[i] >= NIR_GAP):
                            recpr.append(rpr[i])
                   
                recpr.append(pix1) 
            
            
                # print('recpr')   
                # print(recpr)  
                # print()
                
                #  Merge RPs.
                recpnts = recpg + recpr
        
                # Normalize RPs to chromosome length of 1000 pixels.
                rpnorm = [int(element * 1000/pix1) for element in recpnts]
                rpnorm.sort()
                
                # print("rpnorm")
                # print(rpnorm)
                # print()
                
                # Merge rpnorms.
                recombs = [*recombs, *rpnorm]
                
                # print("recombs")
                # print(recombs)
                # print()
        
        # Remove duplicates and sort.
        recombs = [*set(recombs)]
        recombs.sort()
        
        # print('recombs')
        # print(recombs)
        # print
        
        # Initialize final list.
        rpsfinal = []
        
        # Merge RPs within a distance of <= RP_GAP. Remove duplicates and sort.
        for i in range(len(recombs) - 1):
            if recombs[i+1] - recombs[i] <= RP_GAP:
                rpsfinal.append(np.average(recombs[i:i+2]))
                recombs[i+1] = 0
            else:
                rpsfinal.append(recombs[i])
                
        rpsfinal.append(1000)        
        
        rpsfinal = [*set(rpsfinal)]
        rpsfinal.sort()
        
                
        # print('rpsfinal')  
        # print(rpsfinal)
        
        # Merge duplicate RPs
        rpsfin = []
    
        for i in range(len(rpsfinal)-1):
            if (rpsfinal[i+1] - rpsfinal[i] <= DUP_DIST):
                rpsfin.append(np.average(rpsfinal[i:i+2]))
                rpsfinal[i+1] = 0
            else:
                rpsfin.append(rpsfinal[i])
        
        rpsfin.append(1000)        
        
        rpsfin = [*set(rpsfin)]
        rpsfin.sort()
        
        # print('rpsfin')  
        # print(rpsfin)
            
        # Convert pixels to inches for display.
        rpins = [round(element * SCALE_FACTOR,2) for element in rpsfin]
        
        # print('rpins')
        # print(rpins)
        
        # Set column widths.
        rpcols = []
        rpcols.append(rpins[1])
        for i in range(2, len(rpins)):
            rpcols.append(round(rpins[i] - rpins[i-1],2))
            
        # print('rpcols')
        # print(rpcols)
         
        # Implement RP drawing.
        for i, column_width in enumerate(rpcols):  
            ws.column_dimensions[get_column_letter(i+11)].width = column_width
            col_borders_thick(ws, i+11, mps * 8)
            
        col_borders_thick(ws, len(rpcols) + 11, mps * 8 )
        
        names = [item for item in names if item[-1] != 'P']
        
        for w, name in enumerate(names):
            ws.cell(mps * 3  + 4 + w*3, 10).value = name
            ws.cell(mps * 3  + 4 + w*3, 10).alignment = Alignment(horizontal='center')
            for i in range(11, len(rpcols) + 11):
                ws.cell(mps * 3  + 4 + w*3, i).fill = PatternFill("solid", fgColor="FF00FF")
                ws.cell(mps * 3  + 5 + w*3, i).fill = PatternFill("solid", fgColor="0066CC")

            if PHASED_KITS:
                ws.cell(mps * 3  + 4 + w*3, len(rpcols) + 11).value = 'Paternal'
                ws.cell(mps * 3  + 4 + w*3, len(rpcols) + 11).alignment = Alignment(horizontal='center')
                ws.cell(mps * 3  + 5 + w*3, len(rpcols) + 11).value = 'Maternal'
                ws.cell(mps * 3  + 5 + w*3, len(rpcols) + 11).alignment = Alignment(horizontal='center')

        ws.cell(mps * 3  + 4 + len(names)*3, 10).fill = PatternFill("solid", fgColor="FF00FF")
        ws.cell(mps * 3  + 5 + len(names)*3, 10).fill = PatternFill("solid", fgColor="0066CC")
        ws.cell(mps * 3  + 7 + len(names)*3, 10).fill = PatternFill("solid", fgColor="00FFFF")
        ws.cell(mps * 3  + 8 + len(names)*3, 10).fill = PatternFill("solid", fgColor="FFCC00")
        ws.cell(mps * 3  + 10 + len(names)*3, 10).fill = PatternFill("solid", fgColor="FF00FF")
        ws.cell(mps * 3  + 11 + len(names)*3, 10).fill = PatternFill("solid", fgColor="FFCC00")
        ws.cell(mps * 3  + 13 + len(names)*3, 10).fill = PatternFill("solid", fgColor="00FFFF")
        ws.cell(mps * 3  + 14 + len(names)*3, 10).fill = PatternFill("solid", fgColor="0066CC")
        
        
        print('Chromosome {} finished\n'.format(chrom))
                    
def scraperX():
    global src, recapno, src2
    # Prevent Hard Breaks.
    driver.find_element(By.XPATH,'//*[@id="NextPage"]/div[9]/input').click()
      
    # Set cM cutoof
    driver.find_element(By.XPATH, '//*[@id="NextPage"]/div[7]/input').clear()
    driver.find_element(By.XPATH, '//*[@id="NextPage"]/div[7]/input').send_keys(CHROM_X_CUTOFF)

    
    # Send Individuals' GEDmatch ID.
    try:
        driver.find_element(By.XPATH,'//*[@id="NextPage"]/div[1]/input').send_keys(match_pair_ids[q][0])
    except:
        print("There's a problem with Individual1's GEDmatch ID.\n Please check that it was entered correctly.")
        driver.close()
        sys.exit()
    try:
        driver.find_element(By.XPATH,'//*[@id="NextPage"]/div[2]/input').send_keys(match_pair_ids[q][1])
    except:
        print("There's a problem with Individual2's GEDmatch ID.\n Please check that it was entered correctly.")
        driver.close()
        sys.exit() 
            
    # Click "COMPARE". If reCAPTCHA present try manual clicking.
    try:
        driver.find_element(By.XPATH, '//*[@id="NextPage"]/button').click()
    except:
        winsound.Beep(BEEP_TONE, BEEP_TIME)
        driver.maximize_window()
        recapno += 1    
        
        time.sleep(RECAPTCHA_TIMEOUT)

    driver.minimize_window()
    
    results = []
    results2 = []
    
    # Download data
    try:
        results = WebDriverWait(driver, imtime).until(EC.presence_of_all_elements_located((By.XPATH,'//*[@id="Chromosomes"]/div/img')))
        results2 = WebDriverWait(driver, imtime).until(EC.presence_of_all_elements_located((By.XPATH,'//*[@id="Chromosomes"]/img')))
    except :
        # er += 1
        print('Internet failure\n')
        # continue
    finally:
        print( "The data download was successful unless otherwise noted\n")
        
    # Place all source addresses into lists "src" and "src2".
    src = []
    src2 = []
    for img in results:
        src.append(img.get_attribute('src'))
    for img in results2:
        src2.append(img.get_attribute('src'))
    
    return src, src2

def add_bordersX():
    ws = wb['Chr23']
    side = Side(border_style='thick')
    border = Border(left=side)
    for i in range(1, 100):
        for j in range(11, 30):
            ws.cell(i, j).border = border

def fill_sheetsX(matches): 
    for w in range(matches):

        ofs = w * 3

        iname = match_pair_names[w][0] + '-' + match_pair_names[w][1]

        ws = wb['Chr23']
        img = openpyxl.drawing.image.Image(wdir + 'temp{0}{1}.png'.format(w, 0))
        correction_factor = 1000/img.width
        img.width = correction_factor * img.width * xscale
        img.height = img.height * 20
        ws.add_image(img, ws.cell(ofs + 3, 11).coordinate)
        ws.cell(ofs + 3, 10).value = iname
        ws.cell(ofs + 3, 10).alignment = Alignment(horizontal='center')

        
        img = openpyxl.drawing.image.Image(wdir + 'tempa{0}{1}.png'.format(w, 0))
        img.width = correction_factor * img.width * xscale
        img.height = img.height * 15 
        ws.add_image(img, ws.cell(ofs + 4, 11).coordinate)

def sheetX(src, src2):
    
    ws = wb['Chr23']
    
    # Indiviuals' name combination
    iname = match_pair_names[q][0] + '-' + match_pair_names[q][1]

    # Cell formatting
    side = Side(border_style='thin')
    border = Border(left=side, right=side, top=side, bottom=side)
    align = Alignment(horizontal='center')

            
    # Set names column width
    if len(iname) >  ws.column_dimensions['J'].width:
        ws.column_dimensions['J'].width = len(iname) + 2
    
    try:
        # Get chromosome images
        urllib.request.urlretrieve(src[0],"gfg") 
         
    except Exception as err:
        
        print(err, '\n')
        
        # Paste images to sheets
        fill_sheetsX(q)
        
        # Add Bold Borders
        add_bordersX()
                       
        # Delete empty sheets
        delete_sheets()
        
        # Save workbook.
        wb.save(xlname) 

        # Delete temporary images
        delete_images()
        
        # Close browser.        
        driver.close()

        print('\nConnection Failure. Data saved in ' + xlname + '\n')

        sys.exit()
            
    with Image.open('gfg') as im:
          for j in range(1):
            im.seek(im.n_frames // 1 * j)
            try:
                im.save(wdir + 'temp{0}{1}.png'.format(q, 0))
            except:
                print("There's a problem with the working directory.\nPlease check that it was entered correctly.")
                driver.close()
                sys.exit()
                
    print('Chromosome data for chromosme 23 acquired for {0} match\n'.format(iname))

    # Get bar images
    urllib.request.urlretrieve(src2[0],"gfg")
    
    with Image.open('gfg') as im:
          for j in range(1):
            im.seek(im.n_frames // 1 * j)
            try:
                im.save(wdir + 'tempa{0}{1}.png'.format(q, 0))
            except:
                print("There's a problem with the working directory.\nPlease check that it was entered correctly.")
                driver.close()
                sys.exit()  
            
    print('Bar data for chromosme 23 acquired for {0} match\n'.format(iname))
   
       
    # # Paste tables            
            
    rows = driver.find_elements(By.XPATH,' /html/body/div[2]/div[2]/div/table/tbody/tr')
    
    nrows = len(rows)
    # Initialize text variable
    text = ''
    
    # Active appropriate sheet
    ws = wb['Chr23']
                    
    
    if q == 0:
        # format_sheet(ws)
        next_line = 3
        ws.cell(next_line, 2).value = iname

    else:   
        next_line = len(ws['B']) + 2
        ws.cell(next_line, 2).value = iname
  

    for r in range(1, nrows + 1):
        for p in range(1, 9):
            text = driver.find_element(By.XPATH,
                '/html/body/div[2]/div[2]/div/table/tbody/tr[{0}]/td[{1}]'.format(r,p)).text    
            ws.cell(next_line + r , p + 1 ).border = border
            ws.cell(next_line+ r , p + 1 ).alignment = align
            ws.cell(r + next_line, p + 1 ).value = text
                
    print('Table {0}, match {1} pasted\n'.format(1, iname))

if __name__ == '__main__':
    
    global steps, q, count, er, recapno
    
    st = time.time()
    
    # Initialize variables.
    initialize()
        
    print('\nInitialization Completed\n') 
    
    # Log in
    login(url)
  
    #  Initialize counter and error counter
    count = 0
    er = 0
    recapno = 0
    
    if selected_chromosomes[0] != 0:
    
        for q, matches in enumerate(match_pair_names):
            
            # Get src file
            scraper() 
    
            print('Images Found\n')
           
            # Process the src file  
            multisheets(src, src2)
            print('Chromosomes acquired\n')
            elapsed_time()
            # print('Please wait while next match-pair images are acquired\n')
    
            if matches[0][-1] != 'P' and matches[1][-1] != 'P':
                steps += 1
            
                # Go to compare page
                url = 'https://app.gedmatch.com/v_compare1.php'
                driver.get(url)
                time.sleep(INTERNET_WAIT)
                scraper()
                print('Images Found\n')
                
                # Process src file
                multisheets(src, src2)
                print('FIR tables acquired\n')
                
                elapsed_time()
                
            count += 1
    
            if count == len(match_pair_names) and not CHROMOSOME_X:
                print('Download completed. Please wait while finishing up.\n')
            else:    
                print('Please wait while next match-pair images are acquired.\n')
     
            # HIR pass
            steps = 1
            
            # Go to compare page
            driver.get('https://app.gedmatch.com/v_compare1.php')
            time.sleep(INTERNET_WAIT)    
            
        # Paste images to sheets
        fill_sheets(len(match_pair_names))
        
        # Add Bold Borders
        if not AUTO_REC_PNTS:
            add_borders()
           
        # Auto RP
        if AUTO_REC_PNTS and unphflag:
            print('Calculating Recombination Points\n')
            auto_rp(len(match_pair_names), names)
        
    if CHROMOSOME_X:
        
        print('Autosomal Chromsomes completed. Please wait while chromosome 23 is processed.\n')

        # Create worksheet
        ws = wb.create_sheet('Chr23')
        format_sheet(ws)


        for q, matches in enumerate(match_pair_names):
        
            # Go to compare page
            driver.get('https://app.gedmatch.com/vx_compare1.php')
            time.sleep(INTERNET_WAIT)
            
            scraperX()
            
            print('Images Found\n')
           
            # Process the src file
            sheetX(src, src2)
            
            print('Chromosomes acquired\n')
            elapsed_time()
            print('Please wait while next match-pair images are acquired\n')
            
            # Go to compare page
            driver.get('https://app.gedmatch.com/vx_compare1.php')
            time.sleep(INTERNET_WAIT)
        
        print('Download completed. Please wait while finishing up\n')

        # Paste images to sheets
        fill_sheetsX(len(match_pair_names))
        
        # Add Bold Borders
        add_bordersX()
    

    
    # Delete empty sheets
    delete_sheets()
    
    print('There were ' + str(er) + ' errors in this run\n')
    
    print('There were ' + str(recapno) + ' reCAPTCHA events in this run\n')
    
    rtot = round((recapno * RECAPTCHA_TIMEOUT/60),2)
    
    print('Total reCAPTCHA time = ' + str(rtot) + ' minutes\n')

    # Save workbook.
    wb.save(xlname) 

    # Delete temporary images
    delete_images()

    # Close browser.        
    driver.close()
    print('Finished\n')
    
    # Finished warning
    if FINISHED_WARNING:
        for i in range(1, WARNING_BEEPS +1):
            winsound.Beep(BEEP_TONE, BEEP_TIME)
            time.sleep(BEEP_SPACING)

    elapsed_time()