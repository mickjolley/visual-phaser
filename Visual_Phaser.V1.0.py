# -*- coding: utf-8 -*-
"""
Visual_Phaser.V1.0.py performs comparisons between siblings and cousins and
stores the results in a .xlsx file.

© 2026 Mick Jolley (mickj1948@gmail.com)

Optimized for speed using a Hybrid Multiprocessing + Multithreading Architecture.
- Multiprocessing: Distributes chromosome analysis across CPU cores.
- Multithreading: Handles concurrent file I/O (DNA loading) and image generation.
"""
import numpy as np
import pandas as pd
import sys
from itertools import combinations
import os
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils.cell import column_index_from_string as cs
from openpyxl.utils import get_column_letter as cl
import time
import platform
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
import multiprocessing
import threading
from typing import Any, cast
from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image as XLImage

# Re-import all config variables from the external configuration file
from VP_configV1 import (
    FILES_PATH, WORKING_DIRECTORY, MAP_PATH, SIBLINGS, PHASED_FILES,
    EVIL_TWINS, COUSINS, CHROMOSOMES, EXCEL_FILE_NAME, SHOW_NO_MATCHES,
    CHROM_TRUE_SIZE, LINEAR_CHROMOSOME, MERGE_FILES, RESOLUTION,
    AUTO_REC_PNTS, ARP_TOLERANCE, AUTO_RP_ASSIGN, REPAIR_FILES,
    SCALE_FACTOR, HIR_CUTOFF, FIR_CUTOFF, X_HIR_CUTOFF, X_FIR_CUTOFF,
    FIR_TABLES, SCALE_ON, FREEZE_COLUMN, LINUX_FONT_STRING,
    HIR_SNP_MIN, FIR_SNP_MIN, MM_DIST, NO_CALL
)

# Global cache to store loaded DNA data and a lock to manage concurrent access
worker_dna_cache = {}
cache_lock = threading.Lock()


def _looks_like_vcf(file_path):
    if str(file_path).lower().endswith('.vcf'):
        return True

    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as handle:
            for _ in range(20):
                line = handle.readline()
                if not line:
                    break
                stripped = line.strip()
                if stripped.startswith('##fileformat=VCF') or stripped.startswith('#CHROM'):
                    return True
    except OSError:
        return False

    return False


def _parse_vcf_file(file_path):
    header_columns = None
    separator = '\t'

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith('#CHROM'):
                if '\t' in line:
                    header_columns = line.lstrip('#').split('\t')
                    separator = '\t'
                else:
                    header_columns = line.lstrip('#').split()
                    separator = r'\s+'
                break

    if not header_columns:
        return None

    if separator == '\t':
        raw = pd.read_csv(
            file_path,
            sep='\t',
            comment='#',
            header=None,
            names=header_columns,
            dtype=str,
            low_memory=False,
            keep_default_na=False,
        )
    else:
        raw = pd.read_csv(
            file_path,
            sep=r'\s+',
            comment='#',
            header=None,
            names=header_columns,
            dtype=str,
            low_memory=False,
            keep_default_na=False,
            engine='python',
        )

    if raw.empty:
        return None

    chrom_col = _pick_column(raw.columns, ['chrom', 'chromosome'])
    pos_col = _pick_column(raw.columns, ['pos', 'position'])
    id_col = _pick_column(raw.columns, ['id', 'rsid'])
    ref_col = _pick_column(raw.columns, ['ref'])
    alt_col = _pick_column(raw.columns, ['alt'])
    format_col = _pick_column(raw.columns, ['format'])

    if not all([chrom_col, pos_col, ref_col, alt_col]):
        return None

    df = pd.DataFrame({
        'chromosome': raw[chrom_col].astype(str),
        'position': raw[pos_col].astype(str),
        'rsid': raw[id_col].astype(str) if id_col else '',
    })

    # Default to REF/ALT for non-genotype VCF rows.
    ref_series = raw[ref_col].fillna('').astype(str)
    alt_first = raw[alt_col].fillna('').astype(str).str.split(',').str[0]
    df['allele1'] = ref_series
    df['allele2'] = alt_first

    if format_col and len(header_columns) > 9:
        sample_col = header_columns[9]
        format_tokens = raw[format_col].fillna('').astype(str).str.split(':')
        gt_index = format_tokens.apply(lambda toks: toks.index('GT') if 'GT' in toks else -1)
        sample_tokens = raw[sample_col].fillna('').astype(str).str.split(':')

        def decode_gt(gt_idx, sample_vals, ref_val, alt_val):
            if gt_idx < 0 or gt_idx >= len(sample_vals):
                return ref_val, alt_val.split(',')[0] if alt_val else ref_val

            gt = sample_vals[gt_idx].replace('|', '/').strip()
            choices = [ref_val] + alt_val.split(',')
            parts = gt.split('/')

            def pick(part):
                if not part or part == '.':
                    return ''
                if not part.isdigit():
                    return ''
                idx = int(part)
                if 0 <= idx < len(choices):
                    return choices[idx]
                return ''

            a1 = pick(parts[0]) if len(parts) > 0 else ''
            a2 = pick(parts[1]) if len(parts) > 1 else a1
            if not a1:
                a1 = ref_val
            if not a2:
                a2 = a1
            return a1, a2

        decoded = [
            decode_gt(gt_idx, sample_vals, ref_val, alt_val)
            for gt_idx, sample_vals, ref_val, alt_val in zip(
                gt_index.tolist(),
                sample_tokens.tolist(),
                ref_series.tolist(),
                raw[alt_col].fillna('').astype(str).tolist(),
            )
        ]
        df['allele1'] = [pair[0] for pair in decoded]
        df['allele2'] = [pair[1] for pair in decoded]

    # Populate missing IDs with chromosome-position token.
    missing_id = df['rsid'].isin(['', '.', 'nan', 'None'])
    df.loc[missing_id, 'rsid'] = (
        df.loc[missing_id, 'chromosome'].astype(str) + ':' + df.loc[missing_id, 'position'].astype(str)
    )

    return df


def _read_raw_dna_table(file_path):
    if _looks_like_vcf(file_path):
        vcf_df = _parse_vcf_file(file_path)
        if vcf_df is not None:
            return vcf_df

    # Try delimiter auto-detection first, then explicit fallbacks.
    read_attempts = [
        {'sep': None, 'engine': 'python'},
        {'sep': '\t'},
        {'sep': ','},
    ]
    for attempt in read_attempts:
        try:
            df = pd.read_csv(
                file_path,
                skip_blank_lines=True,
                comment='#',
                header=0,
                low_memory=False,
                dtype=str,
                keep_default_na=False,
                **attempt,
            )
            if df is not None and len(df.columns) > 0:
                return df
        except (pd.errors.ParserError, pd.errors.EmptyDataError, UnicodeDecodeError, OSError, ValueError):
            continue
    return None


def _pick_column(columns, aliases):
    normalized = {str(col).strip().lower(): col for col in columns}
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    return None


def _clean_allele(series, no_call_val):
    no_call_token = str(no_call_val).strip().upper()
    cleaned = series.fillna('').astype(str).str.strip().str.upper()
    cleaned = cleaned.str.replace(r'[^A-Z0-9-]', '', regex=True)

    # Common no-call encodings seen across raw DNA exports.
    no_call_aliases = {'', '-', '--', '0', '00', 'N', 'NN', 'NC', 'NOCALL'}
    cleaned = cleaned.where(~cleaned.isin(no_call_aliases), no_call_token)
    cleaned = cleaned.where(cleaned.isin({'A', 'T', 'C', 'G', no_call_token}), no_call_token)
    return cleaned


def agnostic_load_individual_dna(ind, files_path, no_call_val, return_error=False):
    """
    Loads and pre-processes DNA for one individual from any supported raw DNA file.
    This parser is delimiter-agnostic (CSV/TAB) and schema-agnostic for common
    consumer DNA exports (Ancestry/23andMe/MyHeritage/FTDNA-like layouts).
    """
    file_names = os.listdir(files_path)
    candidates = [name for name in file_names if f"{ind}_raw_dna" in name]
    last_error = f"No matching '*{ind}_raw_dna*' file found in FILES_PATH."
    for filname in candidates:
        this_file = os.path.join(files_path, filname)
        try:
            raw = _read_raw_dna_table(this_file)
            if raw is None or raw.empty:
                last_error = f"{filname}: file could not be parsed or produced no rows."
                continue

            # Resolve columns by common aliases first.
            rsid_col = _pick_column(raw.columns, ['rsid', 'rs#', 'snp'])
            chrom_col = _pick_column(raw.columns, ['chromosome', 'chrom', 'chr'])
            pos_col = _pick_column(raw.columns, ['position', 'pos'])
            allele1_col = _pick_column(raw.columns, ['allele1'])
            allele2_col = _pick_column(raw.columns, ['allele2'])
            genotype_col = _pick_column(raw.columns, ['result', 'genotype', 'alleles', 'allele_pair'])

            # Fallback to column count if headers are non-standard.
            if rsid_col is None or chrom_col is None or pos_col is None:
                cols = list(raw.columns)
                if len(cols) >= 4:
                    rsid_col, chrom_col, pos_col = cols[0], cols[1], cols[2]
                    if len(cols) >= 5:
                        allele1_col, allele2_col = cols[3], cols[4]
                    else:
                        genotype_col = cols[3]
                else:
                    last_error = f"{filname}: missing required columns (need rsid/chromosome/position + alleles or genotype)."
                    continue

            df = pd.DataFrame({
                'rsid': raw[rsid_col].astype(str),
                'chromosome': raw[chrom_col].astype(str),
                'position': raw[pos_col].astype(str),
            })

            if allele1_col is not None and allele2_col is not None:
                df['allele1'] = raw[allele1_col]
                df['allele2'] = raw[allele2_col]
            elif genotype_col is not None:
                genotype = raw[genotype_col].fillna('').astype(str).str.strip().str.upper()
                genotype = genotype.str.replace(r'[^A-Z0-9-]', '', regex=True)
                df['allele1'] = genotype.str[0]
                df['allele2'] = genotype.str[1]
            else:
                last_error = f"{filname}: allele columns were not found and no genotype column was available."
                continue

            # Normalize chromosome labels from multiple vendor formats.
            df['chromosome'] = df['chromosome'].str.strip().str.upper().str.replace('CHR', '', regex=False)
            df['chromosome'] = df['chromosome'].replace({'X': '23', 'XY': '23', 'MT': 'M'})
            df = df[~df['chromosome'].isin(['Y', 'M'])]
            df = df[df['chromosome'].str.isnumeric()]
            df['chromosome'] = df['chromosome'].astype(int)

            # Keep only valid genomic positions.
            df['position'] = pd.to_numeric(df['position'], errors='coerce')
            df = df.dropna(subset=['position'])
            df['position'] = df['position'].astype(int)

            df['allele1'] = _clean_allele(df['allele1'], no_call_val)
            df['allele2'] = _clean_allele(df['allele2'], no_call_val)

            if df.empty:
                last_error = f"{filname}: no usable autosomal rows after normalization/filtering."
                continue

            result = (ind, df.sort_values(by='position').reset_index(drop=True))
            if return_error:
                return result + (None,)
            return result
        except Exception as e:
            last_error = f"{filname}: {e}"

    if return_error:
        return ind, None, last_error
    return ind, None

def apply_conditions_vectorized(al1x, al2x, al1y, al2y, no_call_val):
    """
    Determines the match type (HIR, FIR, NIR) for a set of alleles using vectorized operations.
    - Crimson: Fully Identical (Both alleles match on both chromosomes).
    - Limegreen: Half Identical (At least one allele matches).
    - Yellow: No match (Different alleles on both chromosomes).
    """
    cond_nc = (al1x == no_call_val) | (al1y == no_call_val)
    cond_crimson = (al1x == al2x) & (al1y == al2y) & (al1x != al1y)
    cond_limegreen = ((al1x == al1y) & (al2x == al2y)) | ((al1x == al2y) & (al2x == al1y))

    res = np.full(al1x.shape, 'yellow', dtype=object)
    res[cond_limegreen] = 'limegreen'
    res[cond_crimson] = 'crimson'
    res[cond_nc] = 'limegreen' # Treat no-calls as limegreen for continuity
    return res

def scan_genomes_optimized(dm, chrom, hir_cutoff, fir_cutoff, hir_snp_min, fir_snp_min, mm_dist, dmap_positions, dmap_cms):
    """
    Identifies contiguous segments of matching DNA (HIR and FIR).
    Uses a genetic map (min_map.txt) to calculate distances in centiMorgans (cM).
    - dx: Half-identical regions (HIR)
    - ds: Fully-identical regions (FIR)
    """
    matches = dm["match"].values
    positions = dm["position"].values
    length = len(matches)

    dx, ds = [], []
    nmms = 0
    segflag = fflag = False
    stpos = pos = fstpos = fpos = nsnps = fsnps = mmpos = 0

    def get_dcm(start, end):
        """Interpolates cM distance between two genomic positions."""
        stcm = np.interp(start, dmap_positions, dmap_cms)
        fincm = np.interp(end, dmap_positions, dmap_cms)
        return fincm - stcm

    # Iterative scan through the DNA sequence to find segments
    for i in range(length):
        m, p = matches[i], positions[i]
        if not segflag:
            if m in ('yellow', 'limegreen'):
                nsnps, segflag, stpos = 1, True, p
                if m == 'limegreen':
                    fsnps, fstpos, fflag = 1, p, True
        else:
            if m in ('yellow', 'limegreen'):
                nsnps += 1
                pos = p
                if fflag:
                    if m == 'limegreen':
                        fsnps, fpos = fsnps + 1, p
                    else:
                        fflag = False
                        if fsnps > fir_snp_min:
                            dcm = get_dcm(fstpos, fpos)
                            if dcm > fir_cutoff:
                                ds.append({"Chr": chrom, "Start Mb": fstpos, "Finish Mb": fpos, "No. SNPs": fsnps, "Length (cM)": round(dcm, 1)})
                        fsnps = 0
                else:
                    if m == 'limegreen':
                        fsnps, fstpos, fflag = 1, p, True
            else: # m == 'crimson' (No match)
                if fflag:
                    if fsnps > fir_snp_min:
                        dcm = get_dcm(fstpos, fpos)
                        if dcm > fir_cutoff:
                            ds.append({"Chr": chrom, "Start Mb": fstpos, "Finish Mb": fpos, "No. SNPs": fsnps, "Length (cM)": round(dcm, 1)})
                    fflag, fsnps = False, 0

                nmms += 1
                if nmms == 1:
                    mmpos = p
                else:
                    # End segment if mismatches are too close
                    if p - mmpos < mm_dist * 1000:
                        segflag, nmms = False, 0
                        if nsnps > hir_snp_min:
                            dcm = get_dcm(stpos, pos)
                            if dcm > hir_cutoff:
                                dx.append({"Chr": chrom, "Start Mb": stpos, "Finish Mb": pos, "No. SNPs": nsnps, "Length (cM)": round(dcm, 1)})
                        nsnps = 0
                    else:
                        nmms, mmpos = 1, p

    # Capture any segments remaining at the end of the chromosome
    if segflag and nsnps > hir_snp_min:
        dcm = get_dcm(stpos, pos)
        if dcm > hir_cutoff:
            dx.append({"Chr": chrom, "Start Mb": stpos, "Finish Mb": pos, "No. SNPs": nsnps, "Length (cM)": round(dcm, 1)})
    if fflag and fsnps > fir_snp_min:
        dcm = get_dcm(fstpos, fpos)
        if dcm > fir_cutoff:
            ds.append({"Chr": chrom, "Start Mb": fstpos, "Finish Mb": fpos, "No. SNPs": fsnps, "Length (cM)": round(dcm, 1)})

    return pd.DataFrame(dx), pd.DataFrame(ds)

def repair_files_optimized(dm, fir_snp_min, mm_dist):
    """
    Noise reduction: smooths over isolated mismatches or small segments.
    - Fills small gaps in limegreen segments.
    - Reassigns isolated crimson SNPs to yellow.
    """
    matches, positions = dm['match'].values, dm['position'].values
    length = len(matches)
    firs = fir_snp_min // 2
    is_limegreen = (matches == 'limegreen')
    new_matches = matches.copy()

    # Smooth limegreen gaps
    for i in range(firs + 1, length - firs - 1):
        if matches[i] in ('crimson', 'yellow'):
            if np.all(is_limegreen[i-firs : i]) and np.all(is_limegreen[i+1 : i+firs]):
                new_matches[i] = 'limegreen'

    # Identify and reassign isolated crimson SNPs
    crimson_idx = np.where(new_matches == 'crimson')[0]
    if len(crimson_idx) > 0:
        mm_dst = mm_dist * 1000
        for i in range(len(crimson_idx)):
            curr_pos = positions[crimson_idx[i]]
            isolated = True
            if i > 0 and curr_pos - positions[crimson_idx[i-1]] <= mm_dst:
                isolated = False
            if i < len(crimson_idx) - 1 and positions[crimson_idx[i+1]] - curr_pos <= mm_dst:
                isolated = False
            if isolated:
                new_matches[crimson_idx[i]] = 'yellow'

    dm['match'] = new_matches
    return dm

def get_dplot_optimized(q, dtot, dxtot, dstot, pair_name, chrom_true_size, resolution, linear_chromosome, chr_len, siblings):
    """
    Prepares the data for graphical representation (plotting).
    Bins the genetic data into 'pixels' for the output image.
    Determines recombination points by tracking changes in segment types.
    """
    res_val = min(resolution * 1000, len(dtot))
    if chrom_true_size:
        length = dtot['position'].iloc[-1] - dtot['position'].iloc[0]
        div = (len(dtot) * 250000000) // (res_val * length)
    else:
        div = len(dtot) // res_val
    div = max(1, div)

    matches, positions = dtot.iloc[:, q+1].values, dtot['position'].values
    num_bins = len(dtot) // div + 1
    dplot_matches, dplot_positions = np.full(num_bins, 'grey', dtype=object), np.zeros(num_bins)

    # Binning logic: majority/priority voting for bin color
    for b in range(num_bins):
        start, end = b * div, min((b + 1) * div, len(dtot))
        if start >= end:
            break
        counts = Counter(matches[start:end])
        if counts['crimson'] > 0:
            dplot_matches[b] = 'crimson'
        elif counts['yellow'] > counts['grey']:
            dplot_matches[b] = 'yellow'
        elif counts['limegreen'] > counts['grey']:
            dplot_matches[b] = 'limegreen'
        dplot_positions[b] = positions[end-1]

    dplot = pd.DataFrame({'match': dplot_matches, 'position': dplot_positions, 'bar': 'black'})
    # Map HIR (blue) and FIR (orange) segments onto the plot
    for df_tot, color in [(dxtot, 'blue'), (dstot, 'orange')]:
        if len(df_tot) > 0:
            relevant = df_tot[df_tot['pair'] == pair_name]
            for _, row in relevant.iterrows():
                dplot.loc[(dplot['position'] >= row['Start Mb']) & (dplot['position'] <= row['Finish Mb']), 'bar'] = color

    rps, rnames = [], []
    p1, p2 = pair_name.split('-')
    # Track segment changes as potential recombination points (RPs)
    if p1 in siblings and p2 in siblings:
        bar_changes = np.where(dplot['bar'].values[1:] != dplot['bar'].values[:-1])[0] + 1
        for idx in bar_changes:
            rps.append(idx)
            rnames.append(pair_name)

    # Linearize the chromosome if requested (stretches to standard length)
    if linear_chromosome:
        target_res = 10000 if resolution == 10 else 1000
        dplot_final = pd.DataFrame({'match': 'grey', 'bar': 'grey', 'position': np.linspace(0, chr_len, target_res + 1)})
        fracts = (dplot['position'].values / chr_len * target_res).round().astype(int)
        valid = (fracts >= 0) & (fracts <= target_res)
        dplot_final.loc[fracts[valid], 'match'] = dplot['match'].values[valid]
        dplot_final.loc[fracts[valid], 'bar'] = dplot['bar'].values[valid]
        dplot = dplot_final

    return dplot, rps, rnames

def thread_chromosome(chrom, match_pairs, individuals, files_path, map_positions, map_cms, chr_len, siblings, config_params):
    """
    Main worker function for analyzing a single chromosome.
    Orchestrates DNA loading, matching, smoothing, and image preparation.
    Executed in parallel for each chromosome.
    """
    print(f"Analyzing chromosome{chrom}...", flush=True)

    # Step 1: DNA Loading. Uses threading to parallelize disk reads.
    with cache_lock:
        missing_inds = [ind for ind in individuals if ind not in worker_dna_cache]

    if missing_inds:
        with ThreadPoolExecutor(max_workers=min(len(missing_inds), 8)) as threads:
            load_results = threads.map(lambda ind: agnostic_load_individual_dna(ind, files_path, config_params['NO_CALL']), missing_inds)
            with cache_lock:
                for ind, dna_df in load_results:
                    if dna_df is not None:
                        worker_dna_cache[ind] = dna_df

    hir_cutoff = config_params['X_HIR_CUTOFF'] if chrom == 23 else config_params['HIR_CUTOFF']
    fir_cutoff = config_params['X_FIR_CUTOFF'] if chrom == 23 else config_params['FIR_CUTOFF']

    dtot_parts, tables_data = [], []
    dxtot, dstot = pd.DataFrame(), pd.DataFrame()

    # Step 2: Genetic Analysis (CPU-bound)
    for pair in match_pairs:
        pair_name = f"{pair[0]}-{pair[1]}"
        dna1, dna2 = worker_dna_cache.get(pair[0]), worker_dna_cache.get(pair[1])
        if dna1 is None or dna2 is None:
            continue

        # Merge individual DNA data on common genetic markers
        dm = pd.merge(dna1[dna1['chromosome'] == chrom], dna2[dna2['chromosome'] == chrom],
                      on=("rsid", "chromosome", "position"), suffixes=('_1', '_2'))
        if len(dm) == 0:
            continue

        # Vectorized matching and optional repair
        dm["match"] = apply_conditions_vectorized(dm["allele1_1"].values, dm["allele2_1"].values,
                                                 dm["allele1_2"].values, dm["allele2_2"].values,
                                                 config_params['NO_CALL'])

        if config_params['REPAIR_FILES']:
            dm = repair_files_optimized(dm, config_params['FIR_SNP_MIN'], config_params['MM_DIST'])

        # Extract HIR and FIR segments
        dx, ds = scan_genomes_optimized(dm, chrom, hir_cutoff, fir_cutoff, config_params['HIR_SNP_MIN'],
                                       config_params['FIR_SNP_MIN'], config_params['MM_DIST'],
                                       map_positions, map_cms)

        tables_data.append((pair_name, dx, ds))
        dx['pair'], ds['pair'] = pair_name, pair_name
        dxtot = pd.concat([dxtot, dx], ignore_index=True)
        dstot = pd.concat([dstot, ds], ignore_index=True)
        dtot_parts.append(dm[['position', 'match']].rename(columns={'match': pair_name}))

    if not dtot_parts:
        return None

    # Combine all pair matches into a single chromosome master table
    dtot = dtot_parts[0]
    for part in dtot_parts[1:]:
        dtot = pd.merge(dtot, part, on='position', how='outer' if not config_params['MERGE_FILES'] else 'inner')
    dtot = dtot.fillna('grey').sort_values('position').reset_index(drop=True)

    # Step 3: Graphical Preparation and Image Saving. Uses threading for concurrent image saving.
    all_rps, all_rnames, pair_images = [], [], []
    wdir = config_params['WORKING_DIRECTORY'] + "/"
    last_dplot_len = 0

    with ThreadPoolExecutor(max_workers=4) as image_threads:
        for q, pair in enumerate(match_pairs):
            pair_name = f"{pair[0]}-{pair[1]}"
            if pair_name not in dtot.columns:
                continue

            dplot, rps, rnames = get_dplot_optimized(q, dtot, dxtot, dstot, pair_name, config_params['CHROM_TRUE_SIZE'],
                                                   config_params['RESOLUTION'], config_params['LINEAR_CHROMOSOME'],
                                                   chr_len, siblings)
            all_rps.extend(rps)
            all_rnames.extend(rnames)
            last_dplot_len = len(dplot)

            if q == 0 and config_params['SCALE_ON']:
                image_threads.submit(get_scale_img, dplot, chrom, wdir)

            image_threads.submit(get_image_file, dplot, pair_name, chrom, wdir)
            pair_images.append((pair_name, len(dplot)))

    return {
        'chrom': chrom, 'tables': tables_data, 'pair_images': pair_images,
        'arp_info': (all_rps, all_rnames, last_dplot_len),
        'dxtot_pairs': list(dxtot['pair'].unique()) if len(dxtot) > 0 else []
    }

def get_image_file(dplot, pair_name, chrom, wdir):
    """Generates and saves a visual representation of DNA matches for a sibling pair."""
    img = Image.new("RGB", (len(dplot), 35), color="white")
    draw = ImageDraw.Draw(img)
    colors, bars = dplot['match'].values, dplot['bar'].values
    for i in range(len(dplot)):
        draw.line([(i, 0), (i, 19)], fill=colors[i], width=0)  # SNP match row
        draw.line([(i, 20), (i, 34)], fill=bars[i], width=0)   # Segment row (Blue/Orange)
    img.save(f"{wdir}{pair_name} {chrom}.png")

def get_scale_img(dplot, chrom, wdir):
    """Generates a genomic scale image showing positions in Megabases (Mb)."""
    img = Image.new("RGB", (len(dplot) + 30, 35), color="white")
    draw = ImageDraw.Draw(img)
    if platform.system() == 'Windows':
        fnt, fnt1 = ImageFont.truetype("arial.ttf", 13), ImageFont.truetype("arial.ttf", 10)
    elif platform.system() == 'Darwin':
        fnt, fnt1 = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 13), ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 10)
    else:
        fnt, fnt1 = ImageFont.truetype(LINUX_FONT_STRING, 13), ImageFont.truetype(LINUX_FONT_STRING, 10)

    positions = dplot['position'].values
    for i, snp in enumerate(positions):
        if i % 50 == 0:
            draw.text((i, 5), f"{snp / 1000000:0.1f}\n|", font=fnt, fill="black")
        elif i % 5 == 0:
            draw.text((i, 21), '|', font=fnt1, fill="black")
    img.save(f"{wdir}scale {chrom}.png")

def do_arp_main(ws, rps_list, rnames_list, dplot_len, siblings, auto_rp_assign, resolution, scale_factor, arp_tolerance):
    """
    Auto-Recombination Point (ARP) assignment.
    Calculates where chromosomal crossovers likely occurred based on data from all sibling pairs.
    Sets Excel column widths based on recombination block lengths.
    """
    rcomb_list = list(zip(rps_list, rnames_list))
    next_line = find_next_line(ws, 7, 3)
    arp_row = next_line + len(siblings) * 3
    rps_list = sorted(list(set(rps_list)))
    rps_list = [0] + rps_list + [dplot_len - 1]

    rpsf = []
    arp_tol = arp_tolerance * resolution
    for num in rps_list:
        if not any(abs(num - x) < arp_tol for x in rpsf):
            rpsf.append(num)

    rpsf_diff = np.subtract(rpsf[1:], rpsf[:-1])
    for i, rp in enumerate(rpsf_diff):
        ws.column_dimensions[cl(8 + i)].width = rp * scale_factor
        ws.cell(arp_row, 8 + i).value = rp
        ws.cell(arp_row, 8 + i).alignment = Alignment(horizontal='center')
    ws.cell(arp_row, 7).value = 'Column Width'
    ws.cell(arp_row, 7).alignment = Alignment(horizontal='center')

    # Assign blocks to specific siblings if auto-assignment is enabled
    if auto_rp_assign and len(siblings) > 2:
        for i, rp in enumerate(rpsf[1:-1]):
            matches = [rn[1].split('-') for rn in rcomb_list if abs(rn[0] - rp) < arp_tol]
            flat = [s for pair in matches for s in pair]
            if flat:
                keymax = Counter(flat).most_common(1)[0][0]
                cell = ws.cell(arp_row - 3 * len(siblings), 8 + i)
                cell.value, cell.alignment, cell.font = keymax, Alignment(horizontal='right'), Font(bold=True)
    return len(rpsf_diff) + 8

def find_next_line(ws, col, addn):
    """Helper to find the next empty row in a given Excel column."""
    lr = 0
    for i in range(ws.max_row, 0, -1):
        if ws.cell(i, col).value is not None:
            lr = i
            break
    return lr + addn

def paste_tables(ws, dx, ds, pair_name, fir_tables, show_no_matches):
    """Pastes segment data (Start, End, SNPs, cM) into the Excel worksheet."""
    side, align = Side(border_style="thin"), Alignment(horizontal="center")
    border = Border(left=side, right=side, top=side, bottom=side)
    if not show_no_matches and len(dx) == 0:
        return

    def _paste(data, title):
        data.drop('pair', axis=1, inplace=True)
        line = find_next_line(ws, 2, 2)
        if len(data) > 0:
            ws.cell(line, 2).value = title
        for i, col in enumerate(data.columns):
            c = ws.cell(line + 1, 2 + i)
            c.value, c.alignment, c.border = col, align, border
        for i in range(len(data)):
            for j in range(5):
                c = ws.cell(line + 2 + i, 2 + j)
                c.value, c.alignment, c.border = data.iloc[i, j], align, border

    _paste(dx, pair_name)
    if fir_tables and len(ds) > 0:
        _paste(ds, f"{pair_name} FIR Table")

def paste_image_main(fflag, ws, pair_name, chrom, q, wdir, cousins, scale_on, show_no_matches, dxtot_pairs, im_width, dplot_len):
    """Inserts the generated DNA match images into the Excel worksheet."""
    if q == 0 and not cousins and scale_on:
        ws.add_image(XLImage(f"{wdir}scale {chrom}.png"), ws.cell(1, 8).coordinate)
    if not show_no_matches and pair_name not in dxtot_pairs:
        return
    if len(pair_name) > ws.column_dimensions["G"].width:
        ws.column_dimensions["G"].width = len(pair_name) + 2

    img = XLImage(f"{wdir}{pair_name} {chrom}.png")
    if cousins:
        # Scale image if working with cousins in an existing file
        img.width = img.width * im_width / dplot_len
        if fflag[chrom]:
            next_line = find_next_line(ws, 7, 3)
            fflag[chrom] = False
        else:
            next_line = find_next_line(ws, 7, 2)
    else:
        # Normal sibling placement (offset by 2 rows from previous)
        next_line = 3 if q == 0 else find_next_line(ws, 7, 2)

    ws.add_image(img, ws.cell(next_line, 8).coordinate)
    cell = ws.cell(next_line, 7)
    cell.value, cell.alignment = pair_name, Alignment(horizontal="center")

def format_sheet(ws):
    """Sets standard column widths and freezes panes for readability."""
    ws.column_dimensions["A"].width = 1
    for char, w in zip("BCDEFG", [5, 11, 12, 11, 13, 14]):
        ws.column_dimensions[char].width = w

def add_borders(ws, col):
    """Draws vertical borders to demarcate the chromosome visualization area."""
    next_line = find_next_line(ws, 7, 3)
    side = Side(border_style="thick")
    border = Border(left=side)
    for i in range(1, next_line):
        for j in range(8, col + 1):
            ws.cell(i, j).border = border

def add_chroms(ws, col, siblings, auto_rec_pnts):
    """Adds labels and colored background block markers for each sibling."""
    next_line = find_next_line(ws, 7, -len(siblings) * 3 + 1 if auto_rec_pnts else 3)
    for w, name in enumerate(siblings):
        ws.cell(next_line + w * 3, 7).value = name
        ws.cell(next_line + w * 3, 7).alignment = Alignment(horizontal="center")
        for i in range(8, col):
            # Magenta and Neon Green background indicators
            ws.cell(next_line + w * 3, i).fill = PatternFill("solid", fgColor="FF00FF")
            ws.cell(next_line + 1 + w * 3, i).fill = PatternFill("solid", fgColor="98FF00")

    # Legend color markers
    fills = [("FF00FF", 3), ("98FF00", 4), ("00FFFF", 6), ("FFCC00", 7), ("FF00FF", 9), ("FFCC00", 10), ("00FFFF", 12), ("98FF00", 13)]
    for color, row in fills:
        ws.cell(row, col + 1).fill = PatternFill("solid", fgColor=color)

def delete_images(wdir):
    """Clean up: removes temporary .png files generated during the run."""
    for f in os.listdir(wdir):
        if f.endswith(".png"):
            os.remove(os.path.join(wdir, f))


def ensure_visible_worksheet(wb):
    """Guarantee openpyxl can save by keeping at least one visible worksheet."""
    if not wb.worksheets:
        ws = wb.create_sheet("Results")
        ws["A1"] = "No chromosome sheets were generated."
        ws["A2"] = "Check input files and filters in VP_configV1.py."
        return

    visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if not visible_sheets:
        wb.worksheets[0].sheet_state = "visible"

if __name__ == "__main__":
    start_time = time.time()
    FILES_PATH, WORKING_DIRECTORY, MAP_PATH = map(os.path.normpath, [FILES_PATH, WORKING_DIRECTORY, MAP_PATH])
    wdir = WORKING_DIRECTORY + "/"

    # Pre-flight check: ensure all requested individuals have DNA files
    individuals = list(set(SIBLINGS) | set(PHASED_FILES) | set(EVIL_TWINS) | set(COUSINS))
    missing_individuals = [ind for ind in individuals if not any(ind + "_raw" in f for f in os.listdir(FILES_PATH))]
    if missing_individuals:
        print("\n[VP_INPUT_ERROR] DNA file(s) not found.", flush=True)
        print(f"[VP_INPUT_ERROR] Missing individuals: {', '.join(missing_individuals)}.", flush=True)
        print("[VP_INPUT_ERROR] Expected filename pattern includes '<name>_raw'.", flush=True)
        sys.exit(2)

    # Pre-flight check: ensure every configured sibling loads into a usable DataFrame.
    sibling_load_failures = []
    for sibling in SIBLINGS:
        ind, sibling_df, error_text = agnostic_load_individual_dna(sibling, FILES_PATH, NO_CALL, return_error=True)
        if sibling_df is None or sibling_df.empty:
            sibling_load_failures.append((ind, error_text or "Unknown parsing error."))
        else:
            with cache_lock:
                worker_dna_cache[ind] = sibling_df

    if sibling_load_failures:
        print("\n[VP_INPUT_ERROR] One or more SIBLINGS could not be loaded into usable DNA data.", flush=True)
        print("[VP_INPUT_ERROR] Check FILES_PATH and input file format (columns/delimiter/content).", flush=True)
        for ind, reason in sibling_load_failures:
            print(f"[VP_INPUT_ERROR] {ind}: {reason}", flush=True)
        sys.exit(2)

    # Load or create the Excel workbook
    xlname = os.path.join(wdir, f"{EXCEL_FILE_NAME}.xlsx")
    wb = load_workbook(xlname) if COUSINS else Workbook()
    if not COUSINS:
        del wb["Sheet"]

    # Load genetic map (Distance vs genomic position)
    dmap_source = pd.read_csv(os.path.join(MAP_PATH, "min_map.txt"), sep="\t", header=0)
    # Standard chromosome lengths for GRCh37/hg19
    chr_lens = [249250621, 243199373, 198022430, 191154276, 180915260, 171115067, 159138663, 146364022, 141213431, 135534747, 135006516, 133851895, 115169878, 107349540, 102531392, 90354753, 81195210, 78077248, 59128983, 63025520, 48129895, 51304566, 155270560]

    # Setup comparison pairs based on config
    if COUSINS:
        chroms = [c for c in wb.sheetnames if c.startswith("Chr")]
        ws0 = wb[chroms[0]]
        RESOLUTION, ARP_TOLERANCE, mode = ws0.cell(1,1).value, ws0.cell(2,1).value, ws0.cell(4,1).value
        LINEAR_CHROMOSOME, CHROM_TRUE_SIZE, AUTO_REC_PNTS = (mode == 3), (mode == 2), (ws0.cell(5,1).value == 1)
        SIBLINGS = [ws0.cell(i,7).value for i in range(1, ws0.max_row + 1) if ws0.cell(i,7).value and '-' not in str(ws0.cell(i,7).value) and ws0.cell(i,7).value != 'Column Width']
        match_pairs = [(s, c) for s in SIBLINGS for c in COUSINS if s != c]
        CHROMOSOMES = [int(c[3:]) for c in chroms]
    else:
        # Generate combinations for Sibling pairs and Phased files
        match_pairs = list(combinations(SIBLINGS, 2)) + list(combinations(PHASED_FILES, 2))
        for sib in SIBLINGS:
            for etw in EVIL_TWINS:
                if sib not in etw:
                    match_pairs.append((sib, etw))
        if LINEAR_CHROMOSOME:
            AUTO_REC_PNTS = False

    config_params = {
        'HIR_CUTOFF': HIR_CUTOFF, 'FIR_CUTOFF': FIR_CUTOFF, 'X_HIR_CUTOFF': X_HIR_CUTOFF, 'X_FIR_CUTOFF': X_FIR_CUTOFF,
        'HIR_SNP_MIN': HIR_SNP_MIN, 'FIR_SNP_MIN': FIR_SNP_MIN, 'MM_DIST': MM_DIST, 'NO_CALL': NO_CALL,
        'REPAIR_FILES': REPAIR_FILES, 'MERGE_FILES': MERGE_FILES, 'RESOLUTION': RESOLUTION, 'SCALE_ON': SCALE_ON,
        'CHROM_TRUE_SIZE': CHROM_TRUE_SIZE, 'LINEAR_CHROMOSOME': LINEAR_CHROMOSOME, 'WORKING_DIRECTORY': WORKING_DIRECTORY
    }

    chrom_list = CHROMOSOMES if CHROMOSOMES else list(range(1, 24))
    print(f"\nProcessing {len(chrom_list)} chromosomes using Threads and Multiprocessing...\nThis will take a few seconds. Please be patient...\n", flush=True)

    # STEP 4: Parallel Processing Loop
    with ThreadPoolExecutor(max_workers=multiprocessing.cpu_count()) as executor:
        futures = {executor.submit(thread_chromosome, c, match_pairs, individuals, FILES_PATH,
                   dmap_source[dmap_source["Chromosome"] == c].sort_values("Position")["Position"].values,
                   dmap_source[dmap_source["Chromosome"] == c].sort_values("Position")["cM"].values,
                   chr_lens[c-1], SIBLINGS, config_params): c for c in chrom_list}

        chromosome_results = {}

        for future in as_completed(futures):
            res = future.result()
            if not res:
                continue
            chromosome_results[res['chrom']] = res

        for chrom in sorted(chrom_list):
            res = chromosome_results.get(chrom)
            if not res:
                continue
            print(f"Chromosome {chrom} now merging into Excel...", flush=True)

            # Select or create the worksheet for this chromosome
            if COUSINS:
                ws = wb[f"Chr{chrom}"]
                ws_images = cast(list[Any], getattr(ws, "_images", []))
                im_width, last_col = (ws_images[1].width if len(ws_images) > 1 else 100), ws.max_column - 1
            else:
                ws = wb.create_sheet(f"Chr{chrom}")
                ws.cell(1,1).value, ws.cell(2,1).value = RESOLUTION, ARP_TOLERANCE
                ws.cell(4,1).value = 3 if LINEAR_CHROMOSOME else (2 if CHROM_TRUE_SIZE else 1)
                if AUTO_REC_PNTS:
                    ws.cell(5,1).value = 1
                im_width = 0

            ws.freeze_panes = f"{cl(cs(FREEZE_COLUMN)+1)}1"
            format_sheet(ws)

            # Write data tables and images to Excel
            for p_name, dx, ds in res['tables']:
                paste_tables(ws, dx, ds, p_name, FIR_TABLES, SHOW_NO_MATCHES)

            fflag = [True] * 24

            for q, (p_name, dplot_len) in enumerate(res['pair_images']):
                paste_image_main(fflag, ws, p_name, chrom, q, wdir, COUSINS, SCALE_ON, SHOW_NO_MATCHES, res['dxtot_pairs'], im_width, dplot_len)

            # Post-processing: Add Recombination Points and Formatting
            if not COUSINS:
                if AUTO_REC_PNTS:
                    rps_list, rnames_list, dplot_len = res['arp_info']
                    last_col = do_arp_main(ws, rps_list, rnames_list, dplot_len, SIBLINGS, AUTO_RP_ASSIGN, RESOLUTION, SCALE_FACTOR, ARP_TOLERANCE)
                    line_cell = ws.cell(3, 1)
                    if isinstance(line_cell, Cell):
                        line_cell.value = last_col
                    add_chroms(ws, last_col, SIBLINGS, AUTO_REC_PNTS)
                    add_borders(ws, last_col)
                else:
                    add_chroms(ws, 50, SIBLINGS, AUTO_REC_PNTS)
                    add_borders(ws, 50)
            else:
                add_borders(ws, last_col if AUTO_REC_PNTS else 50)

    # Sort worksheets in numeric chromosome order (Chr1, Chr2, ..., Chr23).
    def _sheet_sort_key(title):
        if title.startswith("Chr") and title[3:].isdigit():
            return (0, int(title[3:]))
        return (1, title)

    sorted_titles = sorted((ws.title for ws in wb.worksheets), key=_sheet_sort_key)
    for idx, title in enumerate(sorted_titles):
        target_sheet = wb[title]
        wb.move_sheet(target_sheet, idx - wb.index(target_sheet))

    # Final Save and Cleanup
    ensure_visible_worksheet(wb)
    wb.save(xlname)
    delete_images(wdir)
    total_time = time.time() - start_time
    print(f"\nTotal elapsed time = {total_time//60:.0f} min {total_time % 60: .0f} sec.", flush=True)
    print("\nFinished", flush=True)
