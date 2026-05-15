# -*- coding: utf-8 -*-
"""
Segments_Fill_V1.py auutomatically fills the segments for each sibling.

One chromosome is analyzed at a time.

Reults are saved to an output file and the segment colors for the matches images
reported.

© 2024 Mick Jolley (mickj1948@gmail.com)
"""
import os
import pprint
import sys
from collections import Counter
from copy import copy

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string as cs
from openpyxl.utils import get_column_letter as cl
from PIL import Image

from Segments_Fill_config_V1 import (
    CHROMOSOME,
    INPUT_FILE_NAME,
    OUTPUT_FILE_NAME,
    START_COL,
    START_SIB,
    XLSX_FOLDER,
)


def find_ind(sibs, match_pair_name):
    if "-" not in str(match_pair_name):
        return None
    ind_name = str(match_pair_name).split("-")[1]
    for k in range(len(sibs)):
        if str(sibs[k][0]) == ind_name:
            return k
    return None


def find_ind2(sibs, match_pair_name, ind):
    if "-" not in str(match_pair_name):
        return None
    name = str(match_pair_name).split("-")[0]
    name2 = str(match_pair_name).split("-")[1]
    for k in range(len(sibs)):
        if k != ind:
            if str(sibs[k][0]) == name:
                return k
            if str(sibs[k][0]) == name2:
                return k
    return None


def nir_paste(ws, sibs, ind2, ind, i, colors):
    color1, color2, color3, color4 = colors
    if ind2 is None or ind is None:
        return

    try:
        cell1 = ws.cell(sibs[ind][1], i)
        cell2 = ws.cell(sibs[ind2][1], i)
        if (
            cell1.fill.start_color.index == color1
            and cell2.fill.start_color.index == "FFFFFFFF"
        ):
            cell2.fill = PatternFill(
                start_color=color2, fill_type="solid"
            )

        if (
            cell1.fill.start_color.index == color2
            and cell2.fill.start_color.index == "FFFFFFFF"
        ):
            cell2.fill = PatternFill(
                start_color=color1, fill_type="solid"
            )

        cell3 = ws.cell(sibs[ind][1] + 1, i)
        cell4 = ws.cell(sibs[ind2][1] + 1, i)
        if (
            cell3.fill.start_color.index == color3
            and cell4.fill.start_color.index == "FFFFFFFF"
        ):
            cell4.fill = PatternFill(
                start_color=color4, fill_type="solid"
            )
        if (
            cell3.fill.start_color.index == color4
            and cell4.fill.start_color.index == "FFFFFFFF"
        ):
            cell4.fill = PatternFill(
                start_color=color3, fill_type="solid"
            )
    except (AttributeError, IndexError) as e:
        print(f"Error in nir_paste at col {i}: {e}")


def fir_paste(ws, sibs, ind2, ind, i, colors):
    color1, color2, color3, color4 = colors
    if ind2 is None or ind is None:
        return

    try:
        cell1 = ws.cell(sibs[ind][1], i)
        cell2 = ws.cell(sibs[ind2][1], i)
        if (
            cell1.fill.start_color.index == color1
            and cell2.fill.start_color.index == "FFFFFFFF"
        ):
            cell2.fill = PatternFill(
                start_color=color1, fill_type="solid"
            )
        if (
            cell1.fill.start_color.index == color2
            and cell2.fill.start_color.index == "FFFFFFFF"
        ):
            cell2.fill = PatternFill(
                start_color=color2, fill_type="solid"
            )

        cell3 = ws.cell(sibs[ind][1] + 1, i)
        cell4 = ws.cell(sibs[ind2][1] + 1, i)
        if (
            cell3.fill.start_color.index == color3
            and cell4.fill.start_color.index == "FFFFFFFF"
        ):
            cell4.fill = PatternFill(
                start_color=color3, fill_type="solid"
            )
        if (
            cell3.fill.start_color.index == color4
            and cell4.fill.start_color.index == "FFFFFFFF"
        ):
            cell4.fill = PatternFill(
                start_color=color4, fill_type="solid"
            )
    except (AttributeError, IndexError) as e:
        print(f"Error in fir_paste at col {i}: {e}")


def filter_inds(ws, match_pair_name, sibs, i):
    for w in range(len(sibs)):
        cell = ws.cell(sibs[w][1], i)
        try:
            if cell.fill.start_color.index != "FFFFFFFF":
                if str(sibs[w][0]) in str(match_pair_name):
                    return w
        except AttributeError:
            continue
    return 99


if __name__ == "__main__":
    xlsx_folder = os.path.normpath(XLSX_FOLDER)
    input_file = os.path.join(xlsx_folder,INPUT_FILE_NAME) + ".xlsx"
    output_file = os.path.join(xlsx_folder,OUTPUT_FILE_NAME) + ".xlsx"

    # Load the workbook and worksheet
    wb = load_workbook(input_file)
    ws = wb[f"Chr{CHROMOSOME}"]

    # 1. Robust Metadata Extraction
    # Find the row containing segment widths (Column Width)
    last_row_idx = None
    for i in range(ws.max_row, 0, -1):
        if ws.cell(i, 7).value == "Column Width":
            last_row_idx = i
            break

    last_col = ws.max_column - 2

    # Extract column widths and calculate cumulative sums
    excel_widths = []
    for i in range(8, last_col + 1):
        w = ws.column_dimensions[cl(i)].width
        excel_widths.append(w)

    total_excel_width = sum(excel_widths)
    if total_excel_width == 0:
        print("Error: Total Excel width is 0. Check 'Column Width' row.")
        sys.exit()
    cumulative_excel_widths = np.cumsum(excel_widths)

    # 2. Precise Image-to-Row Mapping
    row_to_match_pair = {}
    for i in range(1, ws.max_row + 1):
        val = ws.cell(i, 7).value
        if val and "-" in str(val) and val != "Column Width":
            # Store with 1-based indexing
            row_to_match_pair[i] = str(val)

    # Dictionary to hold the results for ALL images associated with match_pairs
    all_images_segs = {}
    first_image_pixel_widths = None

    # 3. Per-Image Analysis with Cumulative Scaling
    # The first image (index 0) is ignored as per requirements.

    images_mapped = 0

    for img_idx, img_ref in enumerate(ws._images):
        if img_idx == 0:
            continue

        try:
            anchor = img_ref.anchor
            if not hasattr(anchor, "_from"):
                continue

            # openpyxl anchor rows are 0-indexed, but ws.cell is 1-indexed
            anchor_row = anchor._from.row + 1
            anchor_col = anchor._from.col

            # Images of interest are in column 8 (index 7) or further
            if anchor_col < 7:
                continue

            # Match image to the match_pair in the same row
            match_pair_name = row_to_match_pair.get(anchor_row)
            if not match_pair_name:
                match_pair_name = row_to_match_pair.get(
                    anchor_row + 1
                ) or row_to_match_pair.get(anchor_row - 1)

            if not match_pair_name:
                print(
                    f"Skipping image {img_idx} at row {anchor_row}: No match pair found."
                )
                continue

            images_mapped += 1

        except AttributeError:
            continue

        # Open and process the image
        img = Image.open(img_ref.ref)
        im = np.asarray(img)
        pix_width = im.shape[1]

        # Analyze the first row of the image for color data
        red_channel = im[0, :, 0]
        green_channel = im[0, :, 1]

        current_image_segs = {}
        current_pixel_widths = []

        prev_pixel_boundary = 0
        for idx, cum_width in enumerate(cumulative_excel_widths):
            # Calculate precise boundary using cumulative ratio
            pixel_boundary = round(pix_width * (cum_width / total_excel_width))

            # Ensure we don't exceed image width due to rounding
            pixel_boundary = min(pixel_boundary, pix_width)

            # Current segment range: [prev_pixel_boundary, pixel_boundary)
            segment_width = pixel_boundary - prev_pixel_boundary
            current_pixel_widths.append(segment_width)

            r_count = 0
            g_count = 0
            y_count = 0

            # Analyze pixels within the segment range
            for j in range(
                prev_pixel_boundary + int(segment_width / 4),
                pixel_boundary - int(segment_width / 4),
            ):
                r_val = red_channel[j]
                g_val = green_channel[j]

                # Color detection logic
                if r_val == 255 and g_val == 255:
                    y_count += 1
                elif g_val == 205:
                    g_count += 1
                elif g_val == 20:
                    r_count += 1

            # Determine segment color
            segment_id = idx + 1  # 1-indexed segment ID
            if r_count > 2:
                current_image_segs[segment_id] = "red"
            elif y_count > 0:
                current_image_segs[segment_id] = "yellow"
            else:
                current_image_segs[segment_id] = "green"
            prev_pixel_boundary = pixel_boundary

        all_images_segs[match_pair_name] = current_image_segs

        # Keep track of pixel widths for the first valid image found to update Excel
        if first_image_pixel_widths is None:
            first_image_pixel_widths = current_pixel_widths

    # 4. Excel Update
    if last_row_idx and first_image_pixel_widths:
        for idx, width in enumerate(first_image_pixel_widths):
            ws.cell(last_row_idx, 8 + idx).value = width

    # Output the image results
    pprint.pprint(all_images_segs)


    color1 = ws.cell(3, ws.max_column).fill.start_color.index
    color2 = ws.cell(6, ws.max_column).fill.start_color.index
    color3 = ws.cell(4, ws.max_column).fill.start_color.index
    color4 = ws.cell(7, ws.max_column).fill.start_color.index

    colors = (color1, color2, color3, color4)
    if any(c is None or c == "00000000" for c in colors):
        print(f"Warning: Some palette colors could not be extracted: {colors}")

    siblings = []
    sib_rows = []

    for i in range(ws.max_row, 1, -1):
        cell_val = ws.cell(i, 7).value
        if (
            cell_val is not None
            and str(cell_val) != "Column Width"
            and "-" not in str(cell_val)
        ):
            siblings.append(str(cell_val))
            sib_rows.append(i)

    siblings.reverse()
    sib_rows.reverse()

    sibs = list(zip(siblings, sib_rows, strict=False))

    if not sibs:
        print("Error: No siblings found in column 7.")
        sys.exit()

    rp_row = sibs[0][1] - 1

    if START_SIB == "Auto":
        rp_list = []

        for i in range(8, last_col):
            name = ws.cell(rp_row, i).value
            if name:
                rp_list.append(str(name))

        if not rp_list:
            print("Error: No reference points (RP) found in row above siblings.")
            sys.exit()

        sib = Counter(rp_list).most_common()[-1][0]

        try:
            ind = siblings.index(sib)
        except ValueError:
            print(f"Error: Chosen sibling {sib} not found in siblings list.")
            sys.exit()

        print(f"\nAutomatic choice of START_SIB = {sib}")
    else:
        ind = siblings.index(START_SIB)

    # Blank.
    for i in range(8, last_col + 1):
        for j in range(sibs[0][1], sibs[len(sibs) - 1][1] + 2):
            ws.cell(j, i).fill = PatternFill(start_color="FFFFFFFF", fill_type="solid")

    st = cs(START_COL)

    if st > last_col or st < 8:
        print("\nSTART_COL out of range. Renter.")
        sys.exit()

    # Fill first segment.
    ws.cell(sibs[ind][1], st).fill = PatternFill(start_color=color1, fill_type="solid")
    ws.cell(sibs[ind][1] + 1, st).fill = PatternFill(
        start_color=color3, fill_type="solid"
    )

    # Fill first sib to first RP.
    next_seg = st + 1
    for i in range(next_seg, last_col + 1):
        if ws.cell(rp_row, i - 1).value != sibs[ind][0]:
            ws.cell(sibs[ind][1], i).fill = PatternFill(
                start_color=color1, fill_type="solid"
            )
            ws.cell(sibs[ind][1] + 1, i).fill = PatternFill(
                start_color=color3, fill_type="solid"
            )
        else:
            next_seg = i
            break

    # Fill first sib to second rp.
    for i in range(next_seg, last_col + 1):
        ws.cell(sibs[ind][1], i).fill = PatternFill(
            start_color=color1, fill_type="solid"
        )
        ws.cell(sibs[ind][1] + 1, i).fill = PatternFill(
            start_color=color4, fill_type="solid"
        )
        if ws.cell(rp_row, i).value == sibs[ind][0]:
            break

    # Run through columns and color segments. Repeat three times to clean up
    # loose ends.
    for _ in range(4):
        for i in range(st, last_col + 1):
            for j in range(3, ws.max_row + 1, 2):
                if j in row_to_match_pair:
                    match_pair_name = row_to_match_pair[j]

                    if match_pair_name not in all_images_segs:
                        continue

                    segment_id = i - 7
                    if segment_id not in all_images_segs[match_pair_name]:
                        continue

                    color = all_images_segs[match_pair_name][segment_id]
                    if color == "yellow":
                        continue

                    curr_ind = filter_inds(ws, match_pair_name, sibs, i)
                    if curr_ind == 99:
                        continue

                    if color == "green":
                        ind2 = find_ind2(sibs, match_pair_name, curr_ind)
                        if ind2 is not None:
                            fir_paste(ws, sibs, ind2, curr_ind, i, colors)
                            if i > 8:
                                for k in range(i - 1, 7, -1):
                                    if ws.cell(rp_row, k).value != sibs[ind2][0]:
                                        ws.cell(sibs[ind2][1], k).fill = copy(
                                            ws.cell(sibs[ind2][1], i).fill
                                        )
                                        ws.cell(sibs[ind2][1] + 1, k).fill = copy(
                                            ws.cell(sibs[ind2][1] + 1, i).fill
                                        )
                                    else:
                                        break
                            for k in range(i + 1, last_col + 1):
                                if ws.cell(rp_row, i).value == sibs[ind2][0]:
                                    break
                                ws.cell(sibs[ind2][1], k).fill = copy(
                                    ws.cell(sibs[ind2][1], i).fill
                                )
                                ws.cell(sibs[ind2][1] + 1, k).fill = copy(
                                    ws.cell(sibs[ind2][1] + 1, i).fill
                                )
                                if ws.cell(rp_row, k).value == sibs[ind2][0]:
                                    break

                    elif color == "red":
                        ind2 = find_ind2(sibs, match_pair_name, curr_ind)
                        if ind2 is not None:
                            nir_paste(ws, sibs, ind2, curr_ind, i, colors)
                            if i > 8:
                                for k in range(i - 1, 7, -1):
                                    if ws.cell(rp_row, k).value != sibs[ind2][0]:
                                        ws.cell(sibs[ind2][1], k).fill = copy(
                                            ws.cell(sibs[ind2][1], i).fill
                                        )
                                        ws.cell(sibs[ind2][1] + 1, k).fill = copy(
                                            ws.cell(sibs[ind2][1] + 1, i).fill
                                        )
                                    else:
                                        break
                            for k in range(i + 1, last_col + 1):
                                if ws.cell(rp_row, i).value == sibs[ind2][0]:
                                    break
                                ws.cell(sibs[ind2][1], k).fill = copy(
                                    ws.cell(sibs[ind2][1], i).fill
                                )
                                ws.cell(sibs[ind2][1] + 1, k).fill = copy(
                                    ws.cell(sibs[ind2][1] + 1, i).fill
                                )
                                if ws.cell(rp_row, k).value == sibs[ind2][0]:
                                    break

    # Delete sheets except target sheet.
    if input_file != output_file:
        for sheet_name in wb.sheetnames:
            if sheet_name != f"Chr{CHROMOSOME}":
                # Get the sheet object and remove it
                sheet_to_delete = wb[sheet_name]
                wb.remove(sheet_to_delete)

    ws.sheet_format.defaultRowHeight = 15

    wb.save(output_file)

    print("\nFinished. Results stored in " + output_file)
