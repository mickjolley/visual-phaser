# -*- coding: utf-8 -*-
"""
Parental_Relationship_V4.py determines runs of homogygosity (ROHs) in a subject's
DNA file and stores the results in a .xlsx file and a .csv file. Multiply the
result by four to approximate sibling comparison results.

© 2026 Mick Jolley (mickj1948@gmail.com)

Optimized for speed using a Hybrid Multiprocessing + Multithreading Architecture.
- Multiprocessing: Distributes chromosome analysis across CPU cores.
- Multithreading: Handles concurrent file I/O (DNA loading) and image generation.
"""
import numpy as np
import pandas as pd
import sys
import os
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.cell import column_index_from_string as cs
from openpyxl.utils import get_column_letter as cl
import time
import platform
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
import multiprocessing
import threading
import csv
from openpyxl.drawing.image import Image as XLImage

# Import all config variables from the external configuration file
from PR_config_V4 import (
    DNA_FILES_PATH, WORKING_DIRECTORY, MAP_PATH, SUBJECTS,
    CHROMOSOMES, EXCEL_FILE_NAME, SHOW_NO_ROHS, 
    CHROM_TRUE_SIZE, LINEAR_CHROMOSOME, RESOLUTION,
    ROH_CUTOFF, FREEZE_COLUMN, LINUX_FONT_STRING, ERR_RATE, SNP_MIN, 
    SNP_DENS_MIN   
)

# Global cache to store loaded DNA data and a lock to manage concurrent access
worker_dna_cache = {}
cache_lock = threading.Lock()

def get_subjects_from_path(dna_path):
    """Extracts subject names from filenames in the DNA files path."""
    subjects = []
    if not dna_path or not os.path.isdir(dna_path):
        return []
    try:
        for f in os.listdir(dna_path):
            if '_raw_dna' in f:
                # Assuming format <Name>_raw_dna.txt or similar
                name = f.split('_raw_dna')[0]
                subjects.append(name)
    except Exception as e:
        print(f"Error reading subjects from path: {e}")
    return sorted(list(set(subjects)))

def _read_raw_dna_table(file_path):
                       
    def parsed_table_looks_usable(df):
        return df is not None and len(df.columns) >= 4

    # Try the common raw-DNA delimiters explicitly, then fall back to auto-detection.
    read_attempts = ['\t', ',']
    for sep in read_attempts:
        try:
            df = pd.read_csv(
                file_path,
                skip_blank_lines=True,
                comment='#',
                header=0,
                low_memory=False,
                dtype=str,
                keep_default_na=False,
                sep=sep,
            )
            if parsed_table_looks_usable(df):
                return df
        except (pd.errors.ParserError, pd.errors.EmptyDataError, UnicodeDecodeError, OSError, ValueError):
            continue

    try:
        df = pd.read_csv(
            file_path,
            skip_blank_lines=True,
            comment='#',
            header=0,
            low_memory=False,
            dtype=str,
            keep_default_na=False,
            sep=None,
            engine='python',
        )
        if parsed_table_looks_usable(df):
            return df
    except (pd.errors.ParserError, pd.errors.EmptyDataError, UnicodeDecodeError, OSError, ValueError):
        pass

    return None

def _pick_column(columns, aliases):
    normalized = {str(col).strip().lower(): col for col in columns}
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    return None

def _normalize_dna_dataframe(df, no_call_val='?'):
    # Normalize chromosome labels from multiple vendor formats.
    df['chromosome'] = df['chromosome'].str.strip().str.upper().str.replace('CHR', '', regex=False)
    df['chromosome'] = df['chromosome'].replace({'X': '23', 'XY': '23', 'MT': 'M'})
    df = df[~df['chromosome'].isin(['Y', 'M'])]
    
    # Keep only valid autosomal chromosomes.
    df = df[df['chromosome'].str.isnumeric()]
    df['chromosome'] = df['chromosome'].astype(int)

    # Keep only valid genomic positions.
    df['position'] = pd.to_numeric(df['position'], errors='coerce')
    df = df.dropna(subset=['position'])
    df['position'] = df['position'].astype(int)

    # Filter for valid genetic letters or the designated no-call value
    valid_alleles = {"A", "T", "C", "G"}
    df = df[df["allele1"].isin(valid_alleles) & df["allele2"].isin(valid_alleles)]
        
    return df

def agnostic_load_individual_dna(ind, files_path, no_call_val='?', return_error=False):
    """
    Loads and pre-processes DNA for one individual from any supported raw DNA file.
    This parser is delimiter-agnostic (CSV/TAB) and schema-agnostic for common
    consumer DNA exports (Ancestry/23andMe/MyHeritage/FTDNA-like layouts).
    """
    with cache_lock:
        if ind in worker_dna_cache:
            result = (ind, worker_dna_cache[ind])
            if return_error:
                return result + (None,)
            return result

    if not os.path.isdir(files_path):
        if return_error:
            return ind, None, f"FILES_PATH '{files_path}' is not a directory."
        return ind, None

    file_names = os.listdir(files_path)
    candidates = [name for name in file_names if f"{ind}_raw_dna" in name]
    last_error = f"No matching '*{ind}_raw_dna*' file found in FILES_PATH."
    for filname in candidates:
        this_file = os.path.join(files_path, filname)
        try:
            raw = _read_raw_dna_table(this_file)
            if raw is None or raw.empty:
                last_error = f"{filname}: file could not be parsed or produced no rows."
                continue

            # Resolve columns by common aliases first.
            rsid_col = _pick_column(raw.columns, ['rsid', 'rs#', 'snp'])
            chrom_col = _pick_column(raw.columns, ['chromosome', 'chrom', 'chr'])
            pos_col = _pick_column(raw.columns, ['position', 'pos'])
            allele1_col = _pick_column(raw.columns, ['allele1'])
            allele2_col = _pick_column(raw.columns, ['allele2'])
            genotype_col = _pick_column(raw.columns, ['result', 'genotype', 'alleles', 'allele_pair'])

            # Fallback to column count if headers are non-standard.
            if rsid_col is None or chrom_col is None or pos_col is None:
                cols = list(raw.columns)
                if len(cols) >= 4:
                    rsid_col, chrom_col, pos_col = cols[0], cols[1], cols[2]
                    if len(cols) >= 5:
                        allele1_col, allele2_col = cols[3], cols[4]
                    else:
                        genotype_col = cols[3]
                else:
                    last_error = f"{filname}: missing required columns (need rsid/chromosome/position + alleles or genotype)."
                    continue

            df = pd.DataFrame({
                'rsid': raw[rsid_col].astype(str),
                'chromosome': raw[chrom_col].astype(str),
                'position': raw[pos_col].astype(str),
            })

            if allele1_col is not None and allele2_col is not None:
                df['allele1'] = raw[allele1_col]
                df['allele2'] = raw[allele2_col]
            elif genotype_col is not None:
                genotype = raw[genotype_col].fillna('').astype(str).str.strip().str.upper()
                genotype = genotype.str.replace(r'[^A-Z0-9-]', '', regex=True)
                df['allele1'] = genotype.str[0]
                df['allele2'] = genotype.str[1]
            else:
                last_error = f"{filname}: allele columns were not found and no genotype column was available."
                continue
            
            # Filter for valid genetic letters or the designated no-call value
            valid_alleles = {"A", "T", "C", "G", no_call_val}
            df = df[df["allele1"].isin(valid_alleles) & df["allele2"].isin(valid_alleles)]

            df = _normalize_dna_dataframe(df, no_call_val)

            if df.empty:
                last_error = f"{filname}: no usable autosomal rows after normalization/filtering."
                continue

            print(f"Loaded DNA file successfully: {filname} ({ind})", flush=True)
            result = (ind, df.sort_values(by='position').reset_index(drop=True))
            if return_error:
                return result + (None,)
            return result
        except Exception as e:
            last_error = f"{filname}: {e}"

    if return_error:
        return ind, None, last_error
    return ind, None

def scan_individual_roh(dna_df, chrom, roh_cutoff_mb, err_rate, snp_min, snp_dens_min, dmap_positions, dmap_cms):
    """
    Identifies Runs of Homozygosity (ROH) for a single individual.
    ROHs are contiguous segments where both alleles are identical.
    - roh_cutoff_mb: Minimum segment length in Megabases (Mb).
    Returns (total_cm, list_of_segment_dicts).
    1% error rate allowed. 50 SNP minimum. SNP density >20 SNPs per Mb.    

    """
    if dna_df.empty:
        return 0.0, []
    
    # Homozygous if allele1 == allele2
    is_homo = (dna_df['allele1'] == dna_df['allele2']).values
    positions = dna_df['position'].values
    
    def get_dcm(start, end):
        stcm = np.interp(start, dmap_positions, dmap_cms)
        fincm = np.interp(end, dmap_positions, dmap_cms)
        return fincm - stcm

    total_roh_cm = 0.0
    roh_segments = []
    in_roh = False
    st_idx = 0
    
    for i in range(len(is_homo)):
        if is_homo[i]:
            if not in_roh:
                st_idx = i
                in_roh = True
                ssnp = i
        else:
            if in_roh:
                if i - ssnp < err_rate * 100:
                    en_idx = i - 1
                    st_pos, en_pos = positions[st_idx], positions[en_idx]
                    length_mb = (en_pos - st_pos) / 1000000
                    if length_mb > roh_cutoff_mb:
                        num_snps = en_idx - st_idx + 1
                        if num_snps > snp_min and num_snps/length_mb > snp_dens_min:
                            dcm = get_dcm(st_pos, en_pos)
                            total_roh_cm += dcm
                            roh_segments.append({
                                "Chr": chrom, "Start Mb": round(st_pos / 1000000, 2), 
                                "Finish Mb": round(en_pos / 1000000, 2), 
                                "No. SNPs": num_snps, "Length (cM)": round(dcm, 1),
                                "Length (Mb)": round(length_mb, 1)
                            })
                    in_roh = False
                ssnp = i
                
    if in_roh:
        en_idx = len(is_homo) - 1
        st_pos, en_pos = positions[st_idx], positions[en_idx]
        length_mb = (en_pos - st_pos) / 1000000
        if length_mb > roh_cutoff_mb:
            num_snps = en_idx - st_idx + 1
            if num_snps > snp_min and num_snps/length_mb > snp_dens_min:
                dcm = get_dcm(st_pos, en_pos)
                total_roh_cm += dcm
                roh_segments.append({
                    "Chr": chrom, "Start Mb": round(st_pos / 1000000, 2), 
                    "Finish Mb": round(en_pos / 1000000, 2), 
                    "No. SNPs": num_snps, "Length (cM)": round(dcm, 1),
                    "Length (Mb)": round(length_mb, 1)
                })

    return total_roh_cm, roh_segments

def get_roh_dplot(dna_df, roh_segments_df, chrom_true_size, resolution, linear_chromosome, chr_len):
    """
    Prepares data for ROH graphical representation.
    Supports fixed width (normalized) or variable width (true size).
    """
    res_val = resolution * 1000
    if chrom_true_size:
        # Use standard chromosome length to ensure all images for this Chr have identical width
        num_bins = int((chr_len / 250000000) * res_val)
    else:
        num_bins = res_val
    num_bins = max(1, num_bins)

    # Homozygous = limegreen, Heterozygous = crimson
    is_homo = (dna_df['allele1'] == dna_df['allele2']).values
    matches = np.where(is_homo, 'limegreen', 'crimson')
    positions = dna_df['position'].values
    
    dplot_matches, dplot_positions = np.full(num_bins, 'grey', dtype=object), np.zeros(num_bins)

    indices = np.linspace(0, len(dna_df), num_bins + 1).astype(int)
    for b in range(num_bins):
        start, end = indices[b], indices[b+1]
        if start >= end:
            continue
        bin_matches = matches[start:end]
        counts = Counter(bin_matches)
        if counts['crimson'] > 0:
            dplot_matches[b] = 'crimson'
        elif counts['limegreen'] > 0:
            dplot_matches[b] = 'limegreen'
        else:
            dplot_matches[b] = 'grey'
        dplot_positions[b] = positions[end-1]

    dplot = pd.DataFrame({'match': dplot_matches, 'position': dplot_positions, 'bar': 'black'})
    if not roh_segments_df.empty:
        for _, row in roh_segments_df.iterrows():
            st, en = row['Start Mb'] * 1000000, row['Finish Mb'] * 1000000
            dplot.loc[(dplot['position'] >= st) & (dplot['position'] <= en), 'bar'] = 'orange'
            
    if linear_chromosome:
        target_res = 10000 if resolution == 10 else 1000
        dplot_final = pd.DataFrame({'match': 'grey', 'bar': 'grey', 'position': np.linspace(0, chr_len, target_res + 1)})
        fracts = (dplot['position'].values / chr_len * target_res).round().astype(int)
        valid = (fracts >= 0) & (fracts <= target_res)
        dplot_final.loc[fracts[valid], 'match'] = dplot['match'].values[valid]
        dplot_final.loc[fracts[valid], 'bar'] = dplot['bar'].values[valid]
        dplot = dplot_final

    return dplot

def thread_chromosome(chrom, individuals, files_path, map_positions, map_cms, chr_len, siblings, config_params):
    """
    Main worker function for analyzing a single chromosome.
    Orchestrates DNA loading, matching, smoothing, and image preparation.
    Executed in parallel for each chromosome.
    """
    print(f"Analyzing chromosome{chrom}...", flush=True)

    # Step 1: DNA Loading. Uses threading to parallelize disk reads.
    with cache_lock:
        missing_inds = [ind for ind in individuals if ind not in worker_dna_cache]

    if missing_inds:
        with ThreadPoolExecutor(max_workers=min(len(missing_inds), 8)) as threads:
            load_results = threads.map(lambda ind: agnostic_load_individual_dna(ind, files_path), missing_inds)
            with cache_lock:
                for ind, dna_df in load_results:
                    if dna_df is not None:
                        worker_dna_cache[ind] = dna_df

    current_chrom_dna = {}
    with cache_lock:
        for ind in individuals:
            dna_df = worker_dna_cache.get(ind)
            if dna_df is not None:
                current_chrom_dna[ind] = dna_df[dna_df['chromosome'] == chrom]


    with ThreadPoolExecutor(max_workers=4) as image_threads:
        # ROH Analysis (Must be inside image_threads pool for submitting tasks)
        roh_results = {}
        roh_tables = []
        roh_images = []
        
        roh_cutoff = config_params.get('ROH_CUTOFF', 5)
        err_rate = config_params.get('ERR_RATE', 1)
        snp_dens_min = config_params.get('SNP_DENS_MIN', 20)
        snp_min = config_params.get('SNP_MIN', 50) 
        show_no_rohs = config_params.get('SHOW_NO_ROHS', False)
        
        for ind in individuals:
            dna_df = current_chrom_dna.get(ind)
            if dna_df is not None:
                tcm, segments = scan_individual_roh(dna_df, chrom, roh_cutoff, err_rate, snp_min, snp_dens_min, map_positions, map_cms)
                roh_results[ind] = tcm
                if segments or show_no_rohs:
                    df_segments = pd.DataFrame(segments) if segments else pd.DataFrame()
                    roh_tables.append((ind, df_segments))
                    
                    # Generate ROH image
                    rdplot = get_roh_dplot(dna_df, df_segments, config_params['CHROM_TRUE_SIZE'], 
                                          config_params['RESOLUTION'], config_params['LINEAR_CHROMOSOME'], chr_len)

                    image_threads.submit(get_scale_img, rdplot, chrom, wdir)

                    image_threads.submit(get_image_file, rdplot, ind, chrom, wdir)
                    roh_images.append((ind, len(rdplot)))

    return {
        'chrom': chrom, 
        'roh_results': roh_results,
        'roh_tables': roh_tables,
        'roh_images': roh_images
    }

def get_image_file(dplot, ind, chrom, wdir):
    """Generates and saves a visual representation of ROHs for a subject."""
    img = Image.new("RGB", (len(dplot), 35), color="white")
    draw = ImageDraw.Draw(img)
    colors, bars = dplot['match'].values, dplot['bar'].values
    for i in range(len(dplot)):
        draw.line([(i, 0), (i, 19)], fill=colors[i], width=0) 
        draw.line([(i, 20), (i, 34)], fill=bars[i], width=0)   
    img.save(f"{wdir}{ind} {chrom}.png")

def get_scale_img(dplot, chrom, wdir):   
    """Generates a genomic scale image showing positions in Megabases (Mb)."""
    img = Image.new("RGB", (len(dplot) + 30, 35), color="white")
    draw = ImageDraw.Draw(img)
    if platform.system() == 'Windows':
        fnt, fnt1 = ImageFont.truetype("arial.ttf", 13), ImageFont.truetype("arial.ttf", 10)
    elif platform.system() == 'Darwin':
        fnt, fnt1 = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 13), ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 10)
    else:
        fnt, fnt1 = ImageFont.truetype(LINUX_FONT_STRING, 13), ImageFont.truetype(LINUX_FONT_STRING, 10)

    positions = dplot['position'].values
    for i, snp in enumerate(positions):
        if i % 50 == 0:
            draw.text((i, 5), f"{snp / 1000000:0.1f}\n|", font=fnt, fill="black")
        elif i % 5 == 0:
            draw.text((i, 21), '|', font=fnt1, fill="black")
    img.save(f"{wdir}scale {chrom}.png")


def find_next_line(ws, col, addn):
    """Helper to find the next empty row in a given Excel column."""
    lr = 0
    for i in range(ws.max_row, 0, -1):
        if ws.cell(i, col).value is not None:
            lr = i
            break
    return lr + addn

def paste_tables(ws, dx, ind_name, show_no_rohs, start_row=1):
    """Pastes ROH data (Start, End, SNPs, cM, Mbs) into the Excel worksheet."""
    side, align = Side(border_style="thin"), Alignment(horizontal="center")
    border = Border(left=side, right=side, top=side, bottom=side)
    if not show_no_rohs and len(dx) == 0:
        return start_row

    current_line = start_row

    def _paste(data, title, line):
        if len(data) == 0:
            return line
        
        line += 1 # Add spacing
        ws.cell(line, 2).value = title
        for i, col in enumerate(data.columns):
            c = ws.cell(line + 1, 2 + i)
            c.value, c.alignment, c.border = col, align, border
        for i in range(len(data)):
            for j in range(len(data.columns)):
                c = ws.cell(line + 2 + i, 2 + j)
                c.value, c.alignment, c.border = data.iloc[i, j], align, border
        return line + 2 + len(data)

    current_line = _paste(dx, ind_name, current_line)
    
    return current_line

def paste_image_main(fflag, ws, ind_name, chrom, q, wdir, show_no_rohs, tot_inds, im_width, dplot_len, start_row=1):
    """Inserts the generated ROH images into the Excel worksheet."""
    # Column indices depend on whether we have the extra spacer column G
    text_col = 8 
    img_col = 9 
    text_col_letter = "H" 

    if q == 0:
        ws.add_image(XLImage(f"{wdir}scale {chrom}.png"), ws.cell(1, img_col).coordinate)
    if not show_no_rohs and ind_name not in tot_inds:
        return start_row
    if len(ind_name) > ws.column_dimensions[text_col_letter].width:
        ws.column_dimensions[text_col_letter].width = len(ind_name) + 4

    img = XLImage(f"{wdir}{ind_name} {chrom}.png")
    # Normal sibling placement (offset by 2 rows from previous)
    next_line = max(3, start_row + 2)

    ws.add_image(img, ws.cell(next_line, img_col).coordinate)
    cell = ws.cell(next_line, text_col)
    cell.value, cell.alignment = ind_name, Alignment(horizontal="center")
    
    return next_line

def format_sheet(ws):
    """Sets standard column widths."""
    
    ws.freeze_panes = f"{cl(cs(FREEZE_COLUMN)+1)}1"
    
    chars, widths = "ABCDEFGH", [1, 5, 11, 12, 11, 13, 14, 14]

    for char, w in zip(chars, widths, strict=True):
        ws.column_dimensions[char].width = w

def delete_images(wdir):
    """Clean up: removes temporary .png files generated during the run."""
    for f in os.listdir(wdir):
        if f.endswith(".png"):
            os.remove(os.path.join(wdir, f))

def ensure_visible_worksheet(wb):
    """Guarantee openpyxl can save by keeping at least one visible worksheet."""
    if not wb.worksheets:
        ws = wb.create_sheet("Results")
        ws["A1"] = "No chromosome sheets were generated."
        ws["A2"] = "Check input files and filters in VP_configV1.py."
        return

    visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if not visible_sheets:
        wb.worksheets[0].sheet_state = "visible"

if __name__ == "__main__":
    start_time = time.time()
    
    # 1. Dynamically load INDIVIDUALS from VCF and SUBJECTS from DNA_FILES_PATH
    path_subjects = get_subjects_from_path(DNA_FILES_PATH)
    
    
    # Use ['*'] to explicitly load all subjects from DNA files path.
    if SUBJECTS == ['*']:
        current_subjects = path_subjects
    else:
        current_subjects = SUBJECTS if SUBJECTS else []
    
    # 3. Define the source 
    if current_subjects:
        print(f"\nComputing ROHs for {len(current_subjects)} Subjects from DNA files\n", flush=True)
                
        # Load Path subjects
        subject_load_failures = []
        for subj in current_subjects:
            ind, subj_df, error_text = agnostic_load_individual_dna(subj, DNA_FILES_PATH, return_error=True)
            if subj_df is None or subj_df.empty:
                subject_load_failures.append((ind, error_text or "DNA file not found or unreadable."))
            else:
                with cache_lock:
                    worker_dna_cache[ind] = subj_df
        
        if subject_load_failures:
            print("\n[VP_INPUT_ERROR] One or more SUBJECTS could not be loaded.", flush=True)
            for ind, reason in subject_load_failures:
                print(f"[VP_INPUT_ERROR] {ind}: {reason}", flush=True)
            sys.exit(2)
            
        current_subjects = [s for s in current_subjects if s in worker_dna_cache]
        
        if not current_subjects:
            print("\n[VP_INPUT_ERROR] No individuals or subjects loaded for comparison.", flush=True)
            sys.exit(2)
                    
        # All loaded individuals for processing
        SIBLINGS = current_subjects 
        FILES_PATH = DNA_FILES_PATH # Primarily used for fallback, but cache is pre-loaded now

            
    # Subject-only mode (fallback)
    print(f"\nComputing ROHs for {len(current_subjects)} Subjects from DNA files...\n", flush=True)

    FILES_PATH = DNA_FILES_PATH
    SIBLINGS = current_subjects
    
    # Normalize paths
    FILES_PATH, WORKING_DIRECTORY, MAP_PATH = map(os.path.normpath, [FILES_PATH, WORKING_DIRECTORY, MAP_PATH])
    wdir = WORKING_DIRECTORY + "/"
    
    individuals = list(set(SIBLINGS))

    # Pre-flight check: ensure every configured sibling is in cache and usable.
    # (Handling cases where some might have been missed in group loading above)
    sibling_load_failures = []
    for sibling in SIBLINGS:
        # Check cache first
        if sibling in worker_dna_cache:
            continue
            
        ind, sibling_df, error_text = agnostic_load_individual_dna(sibling, FILES_PATH, return_error=True)
        if sibling_df is None or sibling_df.empty:
            sibling_load_failures.append((ind, error_text or "Missing in cache and could not be loaded."))
        else:
            with cache_lock:
                worker_dna_cache[ind] = sibling_df

    if sibling_load_failures:
        print("\n[VP_INPUT_ERROR] One or more SIBLINGS could not be loaded into usable DNA data.", flush=True)
        for ind, reason in sibling_load_failures:
            print(f"[VP_INPUT_ERROR] {ind}: {reason}", flush=True)
        sys.exit(2)

    # Create the Excel workbook
    xlname = os.path.join(wdir, f"{EXCEL_FILE_NAME}.xlsx")
    wb = Workbook()
    del wb["Sheet"]

    # Load genetic map (Distance vs genomic position)
    dmap_source = pd.read_csv(os.path.join(MAP_PATH, "min_map.txt"), sep="\t", header=0)

    # Standard chromosome lengths for GRCh37/hg19
    chr_lens = [249250621, 243199373, 198022430, 191154276, 180915260, 171115067, 159138663, 146364022, 141213431, 135534747, 135006516, 133851895, 115169878, 107349540, 102531392, 90354753, 81195210, 78077248, 59128983, 63025520, 48129895, 51304566, 155270560]

    config_params = {
        'RESOLUTION': RESOLUTION,
        'CHROM_TRUE_SIZE': CHROM_TRUE_SIZE, 'LINEAR_CHROMOSOME': LINEAR_CHROMOSOME, 
        'WORKING_DIRECTORY': WORKING_DIRECTORY,
        'SHOW_NO_ROHS': SHOW_NO_ROHS,
        'ROH_CUTOFF': ROH_CUTOFF,
        'ERR_RATE': ERR_RATE,
        'SNP_MIN': SNP_MIN,
        'SNP_DENS_MIN': SNP_DENS_MIN
    }

    chrom_list = [int(c) for c in CHROMOSOMES] if CHROMOSOMES else list(range(1, 23))
    print(f"\nProcessing {len(chrom_list)} chromosomes using Threads and Multiprocessing...\nThis will take a few seconds. Please be patient...\n", flush=True)

    # STEP 4: Parallel Processing Loop
    with ThreadPoolExecutor(max_workers=multiprocessing.cpu_count()) as executor:
        futures = {executor.submit(thread_chromosome, c, individuals, FILES_PATH,
                   dmap_source[dmap_source["Chromosome"] == c].sort_values("Position")["Position"].values,
                   dmap_source[dmap_source["Chromosome"] == c].sort_values("Position")["cM"].values,
                   chr_lens[c-1], SIBLINGS, config_params): c for c in chrom_list}

        chromosome_results = {}
        pair_segments = {}
        total_roh_by_ind = {ind: 0.0 for ind in individuals}

        for future in as_completed(futures):
            res = future.result()
            if not res:
                continue
            chromosome_results[res['chrom']] = res
            
            # Aggregate ROH
            if 'roh_results' in res:
                for ind, val in res['roh_results'].items():
                    total_roh_by_ind[ind] += val

        for chrom in sorted(chrom_list):
            res = chromosome_results.get(chrom)
            if not res:
                continue
            
            # Check if there's anything to report ROHs.       
            has_rohs = 'roh_tables' in res and len(res['roh_tables']) > 0
            
            if not has_rohs and not SHOW_NO_ROHS:
                continue

            print(f"Chromosome {chrom} now merging into Excel...", flush=True)

            # Select or create the worksheet for this chromosome
            ws = wb.create_sheet(f"Chr{chrom}")
            im_width = 0

            format_sheet(ws)

            # # Write data tables to Excel
            table_row, image_row = 1, 1

            if 'roh_tables' in res:
                for ind_name, df_roh in res['roh_tables']:
                    table_row = paste_tables(ws, df_roh, f"{ind_name} ROH Table", config_params.get("SHOW_NO_ROHS", False), start_row=table_row)

            fflag = [True] * 24

            # Paste ROH images
            if 'roh_images' in res:
                for q_roh, (roh_name, dplot_len) in enumerate(res['roh_images']):
                    # Offset q by a large number or handle q=0 scale properly
                    # paste_image_main handles scale if q=0. 
                    # If pair_images was empty, first ROH image should handle scale.
                    actual_q = q_roh 
                    image_row = paste_image_main(fflag, ws, roh_name, chrom, actual_q, wdir, config_params.get("SHOW_no_rohs", False), [roh_name], im_width, dplot_len, start_row=image_row)

    # Sort worksheets in numeric chromosome order (Chr1, Chr2, ..., Chr23).
    def _sheet_sort_key(title):
        if title.startswith("Chr") and title[3:].isdigit():
            return (0, int(title[3:]))
        return (1, title)

    sorted_titles = sorted((ws.title for ws in wb.worksheets), key=_sheet_sort_key)
    for idx, title in enumerate(sorted_titles):
        target_sheet = wb[title]
        wb.move_sheet(target_sheet, idx - wb.index(target_sheet))
    
    tot_roh = 0    
    for ind in sorted(total_roh_by_ind.keys()):
        tot_roh = tot_roh + total_roh_by_ind[ind]

    # Final Save and Cleanup
    ensure_visible_worksheet(wb)
    if tot_roh == 0:
        print('\nThere are no ROHs in any individuals. Excel file not saved.')
    else:
        wb.save(xlname)
    delete_images(wdir)
    total_time = time.time() - start_time
    
    print("\nSummary of ROH (Runs of Homozygosity):", flush=True)
    roh_summary_data = []
    for ind in sorted(total_roh_by_ind.keys()):
        print(f"{ind}: Total ROH {total_roh_by_ind[ind]:0.1f} cM", flush=True)
        roh_summary_data.append([ind, round(total_roh_by_ind[ind], 1)])
    
    # Write ROH CSV summary
    if tot_roh > 0:
        roh_csv_path = os.path.join(WORKING_DIRECTORY, f"{EXCEL_FILE_NAME}_ROH.csv")
        with open(roh_csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Individual", "Total ROH cMs"])
            writer.writerows(roh_summary_data)
        print(f"\nROH summary saved to {roh_csv_path}")

    print(f"\nTotal elapsed time = {total_time//60:.0f} min {total_time % 60: .0f} sec.", flush=True)
    print("\nFinished", flush=True)
